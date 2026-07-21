import calendar
import io
import re
from collections.abc import Callable
from dataclasses import dataclass, field
from datetime import date, datetime, time

from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.employee import Employee
from app.models.location import WorkplaceLocation
from app.models.planning import ShiftPlan
from app.schemas.planning import ShiftCreateRequest
from app.services.shift_validation import leave_conflict_reason

# Legacy-Layout (ein flaches "Vorlage"-Sheet mit Mitarbeiter-ID-Spalte) — wird
# beim Import weiterhin erkannt, siehe _parse_legacy_sheet.
TEMPLATE_HEADERS = ["Mitarbeiter-ID", "Datum", "Startzeit", "Endzeit", "Standort", "Mitarbeitername"]
# Layout der neuen Mitarbeiter-Tabs (Mitarbeiter-ID steht im Tab-Kopf, nicht pro Zeile).
TAB_TABLE_HEADERS = ["Datum", "Von", "Bis", "Standort"]
RESERVED_SHEET_NAMES = {"Vorlage", "Mitarbeiterliste", "Standorte"}
_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")

NIGHT_SHIFT_HINT = (
    "Hinweis: Wenn die Endzeit kleiner oder gleich der Startzeit ist (z. B. 22:00 bis 05:00), "
    "erkennt TimeStemple dies automatisch als Nachtschicht und berechnet das Enddatum auf den "
    "folgenden Kalendertag. Es ist keine zusätzliche Eingabe erforderlich."
)

# Analog zum Deckel in ShiftBulkCreateRequest (max. 5000 Schichten auf einmal).
MAX_IMPORT_ROWS = 5000


@dataclass
class ParsedImportRow:
    row_number: int
    sheet_name: str = ""
    employee_id: int | None = None
    employee_name: str | None = None
    location_id: int | None = None
    location_name: str | None = None
    shift_date: date | None = None
    start_time: time | None = None
    end_time: time | None = None
    is_valid: bool = False
    errors: list[str] = field(default_factory=list)


@dataclass
class ImportParseResult:
    total_rows: int
    all_rows: list[ParsedImportRow]
    valid_rows: list[ParsedImportRow]


def _load_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        return openpyxl, Alignment, Font, PatternFill, get_column_letter
    except ImportError as exc:  # pragma: no cover - defensive, mirrors reports.py
        from fastapi import HTTPException, status

        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl nicht installiert. Bitte 'pip install openpyxl' ausführen.",
        ) from exc


def _next_month_dates(today: date) -> list[date]:
    """Alle Kalendertage des auf ``today`` folgenden Monats."""
    year, month = today.year, today.month + 1
    if month > 12:
        month = 1
        year += 1
    days_in_month = calendar.monthrange(year, month)[1]
    return [date(year, month, day) for day in range(1, days_in_month + 1)]


def _safe_sheet_name(raw: str, used: set[str]) -> str:
    """Sanitisiert einen Tabellenblatt-Namen (Excel: max. 31 Zeichen, keine \\/*?:[])."""
    name = _INVALID_SHEET_CHARS.sub("-", raw).strip().strip("'") or "Mitarbeiter"
    name = name[:31]
    base, suffix_i = name, 2
    while name in used or name in RESERVED_SHEET_NAMES:
        suffix = f" ({suffix_i})"
        name = base[: 31 - len(suffix)] + suffix
        suffix_i += 1
    used.add(name)
    return name


def _sheet_ref(sheet_name: str, cell: str = "A1") -> str:
    """Interne Excel-Hyperlink-Referenz auf ein anderes Tabellenblatt."""
    return f"#'{sheet_name.replace(chr(39), chr(39) * 2)}'!{cell}"


def build_template_workbook(db: Session) -> io.BytesIO:
    """
    Erzeugt die herunterladbare Excel-Vorlage: eine Anleitungs-Übersicht, die
    Nachschlage-Sheets "Mitarbeiterliste"/"Standorte" (mit Sprunglinks) sowie
    ein eigenes Tabellenblatt pro aktivem Mitarbeiter mit dem kompletten
    nächsten Kalendermonat vorausgefüllt (Datum + Standardstandort; Von/Bis leer).
    """
    openpyxl, Alignment, Font, PatternFill, get_column_letter = _load_openpyxl()
    from openpyxl.styles import Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    # Farbpalette an das Frontend angelehnt (siehe AdminDashboard.css: #1e3a5f / #2563eb).
    _BRAND_DARK = "1E3A5F"
    _BRAND_BLUE = "2563EB"
    _BRAND_BLUE_TAB = "60A5FA"

    _HDR_FILL = PatternFill("solid", fgColor=_BRAND_DARK)
    _HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    _BANNER_FILL = PatternFill("solid", fgColor=_BRAND_BLUE)
    _BANNER_FONT = Font(bold=True, color="FFFFFF", size=13)
    _LABEL_FONT = Font(bold=True, size=10, color="374151")
    _HINT_FONT = Font(bold=True, color="8A5A00", size=10)
    _HINT_FILL = PatternFill("solid", fgColor="FFF3CD")
    _INFO_FILL = PatternFill("solid", fgColor="EFF6FF")
    _WEEKEND_FILL = PatternFill("solid", fgColor="FFEDD5")
    _BAND_FILL = PatternFill("solid", fgColor="EFF6FF")
    _LINK_FONT = Font(color=_BRAND_BLUE, underline="single", size=10)
    _TITLE_FONT = Font(bold=True, size=16, color=_BRAND_DARK)
    _SUBTITLE_FONT = Font(italic=True, size=10, color="6B7280")
    _THIN = Side(style="thin", color="D0D5DD")
    _BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
    _DATE_FORMAT = "dd.mm.yyyy (ddd)"

    def _header_row(ws_sheet, cols: list[str], row: int = 1) -> None:
        for ci, col in enumerate(cols, 1):
            cell = ws_sheet.cell(row=row, column=ci, value=col)
            cell.font = _HDR_FONT
            cell.fill = _HDR_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER

    def _autofit(ws_sheet) -> None:
        for col in ws_sheet.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws_sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    wb = openpyxl.Workbook()

    employees = db.scalars(
        select(Employee).where(Employee.is_active.is_(True)).order_by(Employee.name)
    ).all()
    locations = db.scalars(select(WorkplaceLocation).order_by(WorkplaceLocation.name)).all()
    locations_by_id = {loc.id: loc.name for loc in locations}
    default_location = locations[0].name if locations else ""

    month_dates = _next_month_dates(date.today())
    month_label = month_dates[0].strftime("%B %Y") if month_dates else ""

    # Sheet-Namen vorab festlegen, damit Mitarbeiterliste und Mitarbeiter-Tabs
    # dieselben (deduplizierten) Namen für die Sprunglinks verwenden.
    used_sheet_names: set[str] = set()
    sheet_name_by_emp_id: dict[int, str] = {
        emp.id: _safe_sheet_name(f"{emp.id} - {emp.name}", used_sheet_names) for emp in employees
    }

    # ── Sheet 1: Vorlage (Anleitung, keine Dateneingabe mehr) ────────────────
    ws1 = wb.active
    ws1.title = "Vorlage"
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = _BRAND_DARK
    ws1.column_dimensions["A"].width = 100
    ws1.row_dimensions[1].height = 34
    title_cell = ws1.cell(row=1, column=1, value="  📋  Schichtplan-Import — Anleitung")
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = _BANNER_FILL
    title_cell.alignment = Alignment(vertical="center")

    info_lines = [
        "",
        "Für jeden aktiven Mitarbeiter gibt es ein eigenes Tabellenblatt (siehe Reiter unten,",
        "oder direkt über die Mitarbeiterliste anklickbar verlinkt),",
        f"bereits vorausgefüllt mit dem kompletten Monat {month_label}.",
        "Bitte im jeweiligen Mitarbeiter-Tab nur Von/Bis (und bei Bedarf Standort) eintragen.",
        "Zeilen ohne Von- und Bis-Zeit werden beim Import automatisch übersprungen.",
    ]
    row_i = 2
    for line in info_lines:
        cell = ws1.cell(row=row_i, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = _SUBTITLE_FONT
        row_i += 1

    row_i += 1
    hint_cell = ws1.cell(row=row_i, column=1, value=NIGHT_SHIFT_HINT)
    hint_cell.font = _HINT_FONT
    hint_cell.fill = _HINT_FILL
    hint_cell.alignment = Alignment(wrap_text=True, vertical="center")
    hint_cell.border = _BORDER
    ws1.row_dimensions[row_i].height = 45

    # ── Sheet 2: Mitarbeiterliste (Nachschlagehilfe mit Sprunglinks) ─────────
    ws2 = wb.create_sheet("Mitarbeiterliste")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = _BRAND_BLUE
    _header_row(ws2, ["ID", "Name", "E-Mail"])
    for i, emp in enumerate(employees):
        r = i + 2
        id_cell = ws2.cell(row=r, column=1, value=emp.id)
        name_cell = ws2.cell(row=r, column=2, value=f"{emp.name}  →")
        mail_cell = ws2.cell(row=r, column=3, value=emp.email)

        name_cell.hyperlink = _sheet_ref(sheet_name_by_emp_id[emp.id])
        name_cell.font = _LINK_FONT

        band = _BAND_FILL if i % 2 == 1 else None
        for cell in (id_cell, name_cell, mail_cell):
            cell.border = _BORDER
            if band is not None:
                cell.fill = band
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = "A1:C1"
    _autofit(ws2)
    ws2.column_dimensions["B"].width = max(ws2.column_dimensions["B"].width or 0, 26)

    # ── Sheet 3: Standorte (Nachschlagehilfe + Dropdown-Quelle) ──────────────
    ws3 = wb.create_sheet("Standorte")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = _BRAND_BLUE
    _header_row(ws3, ["Name"])
    for i, loc in enumerate(locations):
        cell = ws3.cell(row=i + 2, column=1, value=loc.name)
        cell.border = _BORDER
        if i % 2 == 1:
            cell.fill = _BAND_FILL
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = "A1:A1"
    _autofit(ws3)

    # ── Ein Tabellenblatt pro Mitarbeiter ─────────────────────────────────────
    for emp in employees:
        sheet_name = sheet_name_by_emp_id[emp.id]
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = _BRAND_BLUE_TAB
        location_name = locations_by_id.get(emp.assigned_location_id, default_location)

        ws.merge_cells("A1:D1")
        ws.row_dimensions[1].height = 26
        banner = ws.cell(row=1, column=1, value=f"  🗓️  {emp.name} — {month_label}")
        banner.font = _BANNER_FONT
        banner.fill = _BANNER_FILL
        banner.alignment = Alignment(vertical="center")

        info_cells = [
            ws.cell(row=2, column=1, value="Mitarbeiter-ID:"),
            ws.cell(row=2, column=2, value=emp.id),
            ws.cell(row=3, column=1, value="Name:"),
            ws.cell(row=3, column=2, value=emp.name),
            ws.cell(row=4, column=1, value="Standardstandort:"),
            ws.cell(row=4, column=2, value=location_name or "—"),
        ]
        for cell in info_cells:
            cell.fill = _INFO_FILL
            cell.border = _BORDER
        for cell in info_cells[0::2]:
            cell.font = _LABEL_FONT

        back_link = ws.cell(row=2, column=4, value="← Mitarbeiterliste")
        back_link.hyperlink = _sheet_ref("Mitarbeiterliste")
        back_link.font = _LINK_FONT
        back_link.alignment = Alignment(horizontal="right")

        ws.merge_cells("A6:D6")
        hint = ws.cell(row=6, column=1, value=NIGHT_SHIFT_HINT)
        hint.font = _HINT_FONT
        hint.fill = _HINT_FILL
        hint.alignment = Alignment(wrap_text=True, vertical="center")
        hint.border = _BORDER
        ws.row_dimensions[6].height = 45

        _header_row(ws, TAB_TABLE_HEADERS, row=8)

        location_dv = None
        if locations:
            location_dv = DataValidation(
                type="list",
                formula1=f"Standorte!$A$2:$A${len(locations) + 1}",
                allow_blank=True,
            )
            ws.add_data_validation(location_dv)

        center = Alignment(horizontal="center", vertical="center")
        for offset, day in enumerate(month_dates):
            r = 9 + offset
            is_weekend = day.weekday() >= 5  # Samstag=5, Sonntag=6
            date_cell = ws.cell(row=r, column=1, value=day)
            date_cell.number_format = _DATE_FORMAT
            von_cell = ws.cell(row=r, column=2)
            bis_cell = ws.cell(row=r, column=3)
            loc_cell = ws.cell(row=r, column=4, value=location_name)
            if location_dv is not None:
                location_dv.add(loc_cell)
            band = _BAND_FILL if (not is_weekend and offset % 2 == 1) else None
            for cell in (date_cell, von_cell, bis_cell, loc_cell):
                cell.border = _BORDER
                cell.alignment = center
                if is_weekend:
                    cell.fill = _WEEKEND_FILL
                elif band is not None:
                    cell.fill = band

        ws.freeze_panes = "A9"
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = max(len(location_name or ""), 18) + 2

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _parse_int(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _parse_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def _parse_time(value) -> time | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(text, fmt).time()
            except ValueError:
                continue
    return None


def _build_row(
    db: Session,
    row_number: int,
    sheet_name: str,
    employee_id: int | None,
    employee_name: str | None,
    employee_errors: list[str],
    raw_date,
    raw_start,
    raw_end,
    raw_location,
    locations_by_name: dict[str, tuple[int, str]],
    existing_keys: set[tuple[int, date, time]],
    seen_in_file: set[tuple[int, date, time]],
) -> ParsedImportRow:
    """Validiert die Felder einer einzelnen Schicht-Zeile — genutzt sowohl vom
    Legacy-Flat-Sheet-Parser als auch vom Mitarbeiter-Tab-Parser."""
    result = ParsedImportRow(
        row_number=row_number,
        sheet_name=sheet_name,
        employee_id=employee_id,
        employee_name=employee_name,
    )
    errors: list[str] = list(employee_errors)

    if raw_date is None or raw_date == "":
        errors.append("Pflichtfeld fehlt (Datum).")
    if raw_start is None or raw_start == "":
        errors.append("Pflichtfeld fehlt (Startzeit).")
    if raw_end is None or raw_end == "":
        errors.append("Pflichtfeld fehlt (Endzeit).")
    if raw_location is None or raw_location == "":
        errors.append("Pflichtfeld fehlt (Standort).")

    if raw_location not in (None, ""):
        match = locations_by_name.get(str(raw_location).strip().lower())
        if match is None:
            errors.append("Standort nicht gefunden.")
        else:
            result.location_id, result.location_name = match

    shift_date = _parse_date(raw_date)
    if raw_date not in (None, "") and shift_date is None:
        errors.append("Datum ungültig.")
    else:
        result.shift_date = shift_date

    start_time = _parse_time(raw_start)
    if raw_start not in (None, "") and start_time is None:
        errors.append("Startzeit ungültig.")
    else:
        result.start_time = start_time

    end_time = _parse_time(raw_end)
    if raw_end not in (None, "") and end_time is None:
        errors.append("Endzeit ungültig.")
    else:
        result.end_time = end_time

    if not errors and shift_date is not None and start_time is not None and end_time is not None:
        try:
            ShiftCreateRequest(
                employee_id=employee_id,
                location_id=result.location_id,
                shift_date=shift_date,
                start_time=start_time,
                end_time=end_time,
            )
        except ValidationError as exc:
            errors.extend(e["msg"] for e in exc.errors())

    if not errors and employee_id is not None and shift_date is not None and start_time is not None:
        reason = leave_conflict_reason(db, employee_id, shift_date)
        if reason:
            errors.append(reason)

    if not errors and employee_id is not None and shift_date is not None and start_time is not None:
        key = (employee_id, shift_date, start_time)
        if key in existing_keys:
            errors.append("Doppelte Schicht (existiert bereits).")
        elif key in seen_in_file:
            errors.append("Doppelte Schicht (mehrfach in Datei).")
        else:
            seen_in_file.add(key)

    result.errors = errors
    result.is_valid = len(errors) == 0
    return result


def _resolve_legacy_employee(
    raw_employee_id, active_employees: dict[int, str]
) -> tuple[int | None, str | None, list[str]]:
    if raw_employee_id is None or raw_employee_id == "":
        return None, None, ["Pflichtfeld fehlt (Mitarbeiter-ID)."]
    employee_id = _parse_int(raw_employee_id)
    if employee_id is None:
        return None, None, ["Mitarbeiter-ID ungültig."]
    name = active_employees.get(employee_id)
    if name is None:
        return employee_id, None, ["Mitarbeiter nicht gefunden."]
    return employee_id, name, []


def _parse_legacy_sheet(
    db: Session,
    ws,
    active_employees: dict[int, str],
    locations_by_name: dict[str, tuple[int, str]],
    existing_keys: set[tuple[int, date, time]],
    seen_in_file: set[tuple[int, date, time]],
    check_cap: Callable[[], None],
) -> list[ParsedImportRow]:
    """Legacy-Format: ein flaches Sheet mit Mitarbeiter-ID-Spalte (alte Vorlage)."""
    rows_out: list[ParsedImportRow] = []
    row_number = 1  # Zeile 1 = Header
    for raw_row in ws.iter_rows(min_row=2, values_only=True):
        row_number += 1
        if raw_row is None or all(cell is None or cell == "" for cell in raw_row):
            continue
        check_cap()

        raw_employee_id, raw_date, raw_start, raw_end, raw_location = (list(raw_row) + [None] * 5)[:5]
        employee_id, employee_name, employee_errors = _resolve_legacy_employee(raw_employee_id, active_employees)
        rows_out.append(
            _build_row(
                db,
                row_number,
                "Vorlage",
                employee_id,
                employee_name,
                employee_errors,
                raw_date,
                raw_start,
                raw_end,
                raw_location,
                locations_by_name,
                existing_keys,
                seen_in_file,
            )
        )
    return rows_out


def _find_tab_employee_id(rows: list[tuple]) -> int | None:
    """Sucht in den ersten Zeilen eines Tabs nach 'Mitarbeiter-ID:' und liest die ID daneben."""
    for row in rows:
        if not row:
            continue
        label = str(row[0] or "").strip().rstrip(":").lower()
        if label == "mitarbeiter-id" and len(row) > 1:
            return _parse_int(row[1])
    return None


def _find_tab_header_row(rows: list[tuple]) -> int | None:
    """Findet die 1-basierte Zeilennummer der Tabellenkopfzeile (Datum/Von/Bis/Standort)."""
    expected = [h.lower() for h in TAB_TABLE_HEADERS]
    for idx, row in enumerate(rows, start=1):
        if not row:
            continue
        cells = [str(c or "").strip().lower() for c in (list(row) + [None] * 4)[:4]]
        if cells == expected:
            return idx
    return None


def _parse_employee_tab(
    db: Session,
    ws,
    sheet_name: str,
    employee_id: int,
    employee_name: str | None,
    header_row_idx: int,
    locations_by_name: dict[str, tuple[int, str]],
    existing_keys: set[tuple[int, date, time]],
    seen_in_file: set[tuple[int, date, time]],
    check_cap: Callable[[], None],
) -> list[ParsedImportRow]:
    """Neues Format: ein Tabellenblatt pro Mitarbeiter, ID steht im Tab-Kopf."""
    rows_out: list[ParsedImportRow] = []
    employee_errors = [] if employee_name else ["Mitarbeiter nicht gefunden."]

    row_number = header_row_idx
    for raw_row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        row_number += 1
        if raw_row is None:
            continue

        raw_date, raw_start, raw_end, raw_location = (list(raw_row) + [None] * 4)[:4]
        # Datum/Standort sind für den ganzen Monat vorausgefüllt — eine Zeile ohne
        # Von UND Bis bedeutet "kein Dienst an diesem Tag" und wird still übersprungen,
        # nicht als fehlendes Pflichtfeld gemeldet.
        if (raw_start is None or raw_start == "") and (raw_end is None or raw_end == ""):
            continue
        check_cap()

        rows_out.append(
            _build_row(
                db,
                row_number,
                sheet_name,
                employee_id,
                employee_name,
                employee_errors,
                raw_date,
                raw_start,
                raw_end,
                raw_location,
                locations_by_name,
                existing_keys,
                seen_in_file,
            )
        )
    return rows_out


def parse_and_validate_workbook(db: Session, file_bytes: bytes) -> ImportParseResult:
    """
    Parst und validiert die hochgeladene Excel-Datei. Führt keine DB-Writes aus.

    Unterstützt zwei Layouts gleichzeitig:
    - Neu: ein Tabellenblatt pro Mitarbeiter (Mitarbeiter-ID im Tab-Kopf).
    - Legacy: ein flaches "Vorlage"-Sheet mit Mitarbeiter-ID-Spalte pro Zeile
      (Rückwärtskompatibilität für zuvor heruntergeladene Vorlagen).
    """
    openpyxl, *_ = _load_openpyxl()
    from fastapi import HTTPException, status

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Datei konnte nicht gelesen werden: {exc}",
        ) from exc

    # ── Einmaliges Preloading (keine N+1 Queries) ────────────────────────────
    active_employees: dict[int, str] = {
        e.id: e.name for e in db.scalars(select(Employee).where(Employee.is_active.is_(True))).all()
    }
    locations_all = db.scalars(select(WorkplaceLocation)).all()
    locations_by_name: dict[str, tuple[int, str]] = {
        loc.name.strip().lower(): (loc.id, loc.name) for loc in locations_all
    }
    existing_keys: set[tuple[int, date, time]] = {
        (row.employee_id, row.shift_date, row.start_time)
        for row in db.execute(
            select(ShiftPlan.employee_id, ShiftPlan.shift_date, ShiftPlan.start_time)
        ).all()
    }
    seen_in_file: set[tuple[int, date, time]] = set()

    row_count = 0

    def check_cap() -> None:
        nonlocal row_count
        row_count += 1
        if row_count > MAX_IMPORT_ROWS:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Zu viele Zeilen in der Datei (max. {MAX_IMPORT_ROWS}).",
            )

    all_rows: list[ParsedImportRow] = []

    # ── Legacy: flaches "Vorlage"-Sheet mit eigener Mitarbeiter-ID-Spalte ────
    if "Vorlage" in wb.sheetnames:
        ws_legacy = wb["Vorlage"]
        header = next(ws_legacy.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if (
            header
            and str(header[0] or "").strip() == TEMPLATE_HEADERS[0]
            and str(header[1] or "").strip() == TEMPLATE_HEADERS[1]
        ):
            all_rows.extend(
                _parse_legacy_sheet(
                    db, ws_legacy, active_employees, locations_by_name, existing_keys, seen_in_file, check_cap
                )
            )

    # ── Neu: ein Tabellenblatt pro Mitarbeiter ───────────────────────────────
    for sheet_name in wb.sheetnames:
        if sheet_name in RESERVED_SHEET_NAMES:
            continue
        ws = wb[sheet_name]
        preview_rows = list(ws.iter_rows(min_row=1, max_row=15, values_only=True))

        employee_id = _find_tab_employee_id(preview_rows)
        if employee_id is None:
            continue  # kein erkennbarer Mitarbeiter-Tab — ignorieren

        header_row_idx = _find_tab_header_row(preview_rows)
        if header_row_idx is None:
            all_rows.append(
                ParsedImportRow(
                    row_number=1,
                    sheet_name=sheet_name,
                    employee_id=employee_id,
                    employee_name=active_employees.get(employee_id),
                    errors=["Tabellenkopf (Datum/Von/Bis/Standort) im Tabellenblatt nicht gefunden."],
                    is_valid=False,
                )
            )
            continue

        employee_name = active_employees.get(employee_id)
        all_rows.extend(
            _parse_employee_tab(
                db,
                ws,
                sheet_name,
                employee_id,
                employee_name,
                header_row_idx,
                locations_by_name,
                existing_keys,
                seen_in_file,
                check_cap,
            )
        )

    valid_rows = [r for r in all_rows if r.is_valid]
    return ImportParseResult(total_rows=len(all_rows), all_rows=all_rows, valid_rows=valid_rows)
