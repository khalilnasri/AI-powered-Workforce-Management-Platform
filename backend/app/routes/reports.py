"""
Admin Report-Modul: Arbeitszeiten filtern, auswerten und als CSV/Excel exportieren.

GET /admin/reports/attendance        → JSON
GET /admin/reports/attendance.csv    → CSV-Download
GET /admin/reports/summary           → Chart-Daten (KPIs pro Standort/Monat/Mitarbeiter)
GET /admin/reports/excel             → Excel-Download (.xlsx, 4 Sheets)
"""

from __future__ import annotations

import csv
import io
import logging
import traceback
from datetime import UTC, date, datetime, timedelta
from zoneinfo import ZoneInfo

logger = logging.getLogger(__name__)

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import case, func, select
from sqlalchemy.orm import Session

from app.auth.admin_deps import require_admin
from app.config.database import get_db
from app.models.attendance import Attendance
from app.models.employee import Employee
from app.models.employee_work_location import EmployeeWorkLocation
from app.models.location import WorkplaceLocation
from app.models.work_session import WorkSession
from app.utils.distance import haversine_meters
from app.schemas.admin import (
    AttendanceReportResponse,
    EmployeeReportRow,
    EmployeeSollIstItem,
    LocationHoursItem,
    MonthlyHoursItem,
    ReportSession,
    ReportSummaryResponse,
    ReportV2EmployeeRow,
    ReportV2KPIs,
    ReportV2LocationRow,
    ReportV2PeriodSummary,
    ReportV2Response,
    ReportV2SessionRow,
    ReportV2TrendRow,
    StatusDistributionItem,
)
from app.services.employment_hours import resolved_month_target_hours
from app.services.work_session_stats import get_ws_status_by_checkin_id

router = APIRouter(prefix="/admin/reports", tags=["reports"])

_BERLIN = ZoneInfo("Europe/Berlin")
_WEEKDAYS_DE = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag", "Sonntag"]


# ── Hilfsfunktionen ───────────────────────────────────────────────────────────

def _ensure_utc(dt: datetime) -> datetime:
    """Stellt sicher, dass ein datetime timezone-aware (UTC) ist."""
    if dt.tzinfo is None:
        return dt.replace(tzinfo=UTC)
    return dt.astimezone(UTC)


def _parse_date(date_str: str, end_of_day: bool = False) -> datetime:
    """
    Wandelt 'YYYY-MM-DD' in ein UTC-datetime um.
    end_of_day=True → 23:59:59 desselben Tages (inklusiv).
    """
    try:
        dt = datetime.fromisoformat(date_str).replace(tzinfo=UTC)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Ungültiges Datumsformat: '{date_str}'. Erwartet: YYYY-MM-DD",
        )
    if end_of_day:
        dt = dt + timedelta(days=1) - timedelta(seconds=1)
    return dt


def _build_sessions(
    logs: list[Attendance],
    ws_by_checkin_id: dict[int, str] | None = None,
) -> tuple[list[ReportSession], int]:
    """
    Paart checkin/checkout Zeilen zu Schichten.
    ws_by_checkin_id: checkin_log_id → work_session_status (optional).
    """
    now = datetime.now(UTC)
    sessions: list[ReportSession] = []
    total_seconds = 0
    pending_checkin: datetime | None = None
    pending_checkin_id: int | None = None

    for log in logs:
        if log.log_type == "checkin":
            pending_checkin = _ensure_utc(log.created_at)
            pending_checkin_id = log.id
        elif log.log_type == "checkout" and pending_checkin is not None:
            checkout_at = _ensure_utc(log.created_at)
            secs = max(0, int((checkout_at - pending_checkin).total_seconds()))
            total_seconds += secs
            ws_status = (ws_by_checkin_id or {}).get(pending_checkin_id) if pending_checkin_id else None
            sessions.append(ReportSession(
                checkin=pending_checkin,
                checkout=checkout_at,
                duration_seconds=secs,
                duration_hours=round(secs / 3600, 2),
                status="closed",
                work_session_status=ws_status,
            ))
            pending_checkin = None
            pending_checkin_id = None

    # Offene Schicht (noch eingestempelt)
    if pending_checkin is not None:
        secs = max(0, int((now - pending_checkin).total_seconds()))
        total_seconds += secs
        ws_status = (ws_by_checkin_id or {}).get(pending_checkin_id) if pending_checkin_id else None
        sessions.append(ReportSession(
            checkin=pending_checkin,
            checkout=None,
            duration_seconds=secs,
            duration_hours=round(secs / 3600, 2),
            status="open",
            work_session_status=ws_status,
        ))

    return sessions, total_seconds


def _load_report_data(
    db: Session,
    employee_id: int | None,
    start_date: str | None,
    end_date: str | None,
) -> AttendanceReportResponse:
    """
    Kernlogik: Lädt Attendance-Logs, filtert nach Datum und Employee,
    paart Schichten und berechnet Gesamtzeiten.
    Offizielle Stunden = WorkSessions mit status 'approved' oder 'corrected'.
    """
    # Mitarbeiter laden (alle oder einen bestimmten)
    emp_stmt = select(Employee).order_by(Employee.id)
    if employee_id is not None:
        emp_stmt = emp_stmt.where(Employee.id == employee_id)
    employees = db.scalars(emp_stmt).all()

    if employee_id is not None and not employees:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Mitarbeiter nicht gefunden.")

    # Datumsgrenzen berechnen
    start_dt = _parse_date(start_date) if start_date else None
    end_dt = _parse_date(end_date, end_of_day=True) if end_date else None

    # Alle WorkSessions vorladen (für approved_hours + work_session_status)
    ws_stmt = select(WorkSession)
    if employee_id is not None:
        ws_stmt = ws_stmt.where(WorkSession.employee_id == employee_id)
    if start_dt:
        ws_stmt = ws_stmt.where(WorkSession.checkin_time >= start_dt)
    if end_dt:
        ws_stmt = ws_stmt.where(WorkSession.checkin_time <= end_dt)

    approved_secs_by_emp: dict[int, int] = {}
    # checkin_log_id → status (für alle Mitarbeiter in diesem Report)
    ws_status_by_checkin_id: dict[int, str] = {}
    for ws in db.scalars(ws_stmt).all():
        if ws.status in ("approved", "corrected"):
            approved_secs_by_emp[ws.employee_id] = (
                approved_secs_by_emp.get(ws.employee_id, 0) + ws.duration_seconds
            )
        if ws.checkin_log_id is not None:
            ws_status_by_checkin_id[ws.checkin_log_id] = ws.status

    employee_rows: list[EmployeeReportRow] = []

    for emp in employees:
        # Attendance-Logs für diesen Mitarbeiter laden (chronologisch)
        att_stmt = (
            select(Attendance)
            .where(Attendance.employee_id == emp.id)
            .order_by(Attendance.created_at.asc(), Attendance.id.asc())
        )
        if start_dt:
            att_stmt = att_stmt.where(Attendance.created_at >= start_dt)
        if end_dt:
            att_stmt = att_stmt.where(Attendance.created_at <= end_dt)

        logs = list(db.scalars(att_stmt).all())
        sessions, total_seconds = _build_sessions(logs, ws_status_by_checkin_id)

        approved_secs = approved_secs_by_emp.get(emp.id, 0)

        employee_rows.append(EmployeeReportRow(
            employee_id=emp.id,
            employee_name=emp.name,
            employee_email=emp.email,
            total_seconds=total_seconds,
            total_hours=round(total_seconds / 3600, 2),
            sessions=sessions,
            approved_seconds=approved_secs,
            approved_hours=round(approved_secs / 3600, 2),
        ))

    grand_total_seconds   = sum(r.total_seconds for r in employee_rows)
    approved_total_seconds = sum(r.approved_seconds for r in employee_rows)
    session_count = sum(len(r.sessions) for r in employee_rows)

    return AttendanceReportResponse(
        employees=employee_rows,
        total_seconds=grand_total_seconds,
        total_hours=round(grand_total_seconds / 3600, 2),
        session_count=session_count,
        start_date=start_date,
        end_date=end_date,
        approved_total_seconds=approved_total_seconds,
        approved_total_hours=round(approved_total_seconds / 3600, 2),
    )


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/attendance", response_model=AttendanceReportResponse)
def attendance_report(
    employee_id: int | None = Query(default=None, description="Nur diesen Mitarbeiter auswerten"),
    start_date: str | None = Query(default=None, description="Startdatum YYYY-MM-DD"),
    end_date: str | None = Query(default=None, description="Enddatum YYYY-MM-DD"),
    db: Session = Depends(get_db),
    _: Employee = Depends(require_admin),
):
    """
    JSON-Report: Arbeitszeiten aller (oder eines) Mitarbeiter(s), optional nach Datum gefiltert.
    """
    return _load_report_data(db, employee_id, start_date, end_date)


@router.get("/attendance.csv")
def attendance_report_csv(
    employee_id: int | None = Query(default=None),
    start_date: str | None = Query(default=None),
    end_date: str | None = Query(default=None),
    db: Session = Depends(get_db),
    _: Employee = Depends(require_admin),
):
    """
    CSV-Export: Jede Schicht als eigene Zeile.

    Spalten: Mitarbeiter, E-Mail, Check-in, Check-out, Dauer (Std.), Status
    """
    report = _load_report_data(db, employee_id, start_date, end_date)

    # CSV in-memory aufbauen
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")

    # Kopfzeile
    writer.writerow(["Mitarbeiter", "E-Mail", "Check-in", "Check-out", "Dauer (Std.)", "Status", "Genehmigung"])

    _approval_labels = {
        "approved":  "Genehmigt",
        "corrected": "Korrigiert",
        "rejected":  "Abgelehnt",
        "pending":   "Ausstehend",
    }

    for emp_row in report.employees:
        if not emp_row.sessions:
            writer.writerow([emp_row.employee_name, emp_row.employee_email, "", "", "", "", ""])
            continue
        for session in emp_row.sessions:
            checkin_str = session.checkin.strftime("%d.%m.%Y %H:%M:%S")
            checkout_str = session.checkout.strftime("%d.%m.%Y %H:%M:%S") if session.checkout else "noch offen"
            approval_label = _approval_labels.get(session.work_session_status or "", "—")
            writer.writerow([
                emp_row.employee_name,
                emp_row.employee_email,
                checkin_str,
                checkout_str,
                str(session.duration_hours).replace(".", ","),
                "offen" if session.status == "open" else "geschlossen",
                approval_label,
            ])

    # Dateiname mit Zeitstempel
    now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"time_stemple_report_{now_str}.csv"

    # Bytes mit BOM für Excel-Kompatibilität
    csv_bytes = ("\ufeff" + output.getvalue()).encode("utf-8")

    return StreamingResponse(
        iter([csv_bytes]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Summary (Chart-Daten) ─────────────────────────────────────────────────────

def _month_bounds_utc(year: int, month: int) -> tuple[datetime, datetime]:
    start = datetime(year, month, 1, 0, 0, 0, tzinfo=_BERLIN)
    if month == 12:
        end = datetime(year + 1, 1, 1, 0, 0, 0, tzinfo=_BERLIN)
    else:
        end = datetime(year, month + 1, 1, 0, 0, 0, tzinfo=_BERLIN)
    return start.astimezone(UTC), end.astimezone(UTC)


def _emp_ids_for_location(db: Session, location_id: int) -> set[int]:
    m2m = {
        r.employee_id
        for r in db.execute(
            select(EmployeeWorkLocation.employee_id).where(
                EmployeeWorkLocation.location_id == location_id
            )
        ).all()
    }
    legacy = {
        e.id
        for e in db.scalars(
            select(Employee).where(Employee.assigned_location_id == location_id)
        ).all()
    }
    return m2m | legacy


@router.get("/summary", response_model=ReportSummaryResponse)
def reports_summary(
    month: int | None = Query(default=None, ge=1, le=12),
    year: int | None = Query(default=None, ge=2020, le=2100),
    location_id: int | None = Query(default=None),
    employee_id: int | None = Query(default=None),
    db: Session = Depends(get_db),
    _: Employee = Depends(require_admin),
):
    now_berlin = datetime.now(_BERLIN)
    target_year = year or now_berlin.year
    target_month = month or now_berlin.month

    # ── Mitarbeiter-Filter zusammenbauen ─────────────────────────────────────
    emp_ids: set[int] | None = None
    if location_id is not None:
        emp_ids = _emp_ids_for_location(db, location_id)
    if employee_id is not None:
        emp_ids = (emp_ids & {employee_id}) if emp_ids is not None else {employee_id}

    def _apply_emp_filter(stmt):
        if emp_ids is not None:
            return stmt.where(WorkSession.employee_id.in_(list(emp_ids) or [-1]))
        return stmt

    # ── Zeitgrenzen für soll_vs_ist / status_distribution ───────────────────
    start_utc, end_utc = _month_bounds_utc(target_year, target_month)

    # ── WorkSessions für den Zielmonat laden ─────────────────────────────────
    month_ws = db.scalars(
        _apply_emp_filter(
            select(WorkSession)
            .where(WorkSession.checkin_time >= start_utc)
            .where(WorkSession.checkin_time < end_utc)
        )
    ).all()

    # ── 1. Stunden pro Standort ───────────────────────────────────────────────
    loc_map = {loc.id: loc.name for loc in db.scalars(select(WorkplaceLocation)).all()}
    emp_to_loc: dict[int, str] = {}
    for row in db.execute(
        select(EmployeeWorkLocation.employee_id, EmployeeWorkLocation.location_id)
        .order_by(EmployeeWorkLocation.location_id)
    ).all():
        if row.employee_id not in emp_to_loc:
            emp_to_loc[row.employee_id] = loc_map.get(row.location_id, "Unbekannt")
    for emp in db.scalars(
        select(Employee).where(Employee.assigned_location_id.isnot(None))
    ).all():
        if emp.id not in emp_to_loc and emp.assigned_location_id:
            emp_to_loc[emp.id] = loc_map.get(emp.assigned_location_id, "Unbekannt")

    loc_hours: dict[str, float] = {}
    for ws in month_ws:
        if ws.status not in ("approved", "corrected"):
            continue
        loc_name = emp_to_loc.get(ws.employee_id, "Kein Standort")
        loc_hours[loc_name] = loc_hours.get(loc_name, 0.0) + ws.duration_seconds / 3600

    hours_by_location = sorted(
        [LocationHoursItem(location_name=k, official_hours=round(v, 2)) for k, v in loc_hours.items()],
        key=lambda x: x.official_hours,
        reverse=True,
    )

    # ── 2. Monatlicher Trend (letzte 6 Monate) ────────────────────────────────
    hours_by_month = []
    for i in range(5, -1, -1):
        m = now_berlin.month - i
        y = now_berlin.year
        while m < 1:
            m += 12
            y -= 1
        s_utc, e_utc = _month_bounds_utc(y, m)
        off_sec = db.scalar(
            _apply_emp_filter(
                select(
                    func.coalesce(
                        func.sum(
                            case(
                                (WorkSession.status.in_(("approved", "corrected")), WorkSession.duration_seconds),
                                else_=0,
                            )
                        ),
                        0,
                    )
                )
                .where(WorkSession.checkin_time >= s_utc)
                .where(WorkSession.checkin_time < e_utc)
            )
        ) or 0
        hours_by_month.append(
            MonthlyHoursItem(month=f"{y:04d}-{m:02d}", official_hours=round(off_sec / 3600, 2))
        )

    # ── 3. Soll vs. Ist pro Mitarbeiter ──────────────────────────────────────
    emp_stmt = select(Employee).order_by(Employee.name)
    if emp_ids is not None:
        emp_stmt = emp_stmt.where(Employee.id.in_(list(emp_ids) or [-1]))
    employees_list = db.scalars(emp_stmt).all()

    soll_vs_ist: list[EmployeeSollIstItem] = []
    for emp in employees_list:
        off_sec = db.scalar(
            select(
                func.coalesce(
                    func.sum(
                        case(
                            (WorkSession.status.in_(("approved", "corrected")), WorkSession.duration_seconds),
                            else_=0,
                        )
                    ),
                    0,
                )
            )
            .where(WorkSession.employee_id == emp.id)
            .where(WorkSession.checkin_time >= start_utc)
            .where(WorkSession.checkin_time < end_utc)
        ) or 0
        soll_vs_ist.append(
            EmployeeSollIstItem(
                employee_name=emp.name,
                target_hours=float(resolved_month_target_hours(emp)),
                official_hours=round(off_sec / 3600, 2),
            )
        )

    # ── 4. Status-Verteilung ──────────────────────────────────────────────────
    status_counts: dict[str, int] = {"approved": 0, "corrected": 0, "pending": 0, "rejected": 0}
    for ws in month_ws:
        if ws.status in status_counts:
            status_counts[ws.status] += 1

    return ReportSummaryResponse(
        hours_by_location=hours_by_location,
        hours_by_month=hours_by_month,
        soll_vs_ist=soll_vs_ist,
        status_distribution=StatusDistributionItem(**status_counts),
    )


# ── Excel-Export ──────────────────────────────────────────────────────────────

@router.get("/excel")
def reports_excel(
    month: int | None = Query(default=None, ge=1, le=12),
    year: int | None = Query(default=None, ge=2020, le=2100),
    location_id: int | None = Query(default=None),
    employee_id: int | None = Query(default=None),
    db: Session = Depends(get_db),
    _: Employee = Depends(require_admin),
):
    """Excel-Export (.xlsx) mit 4 Sheets: Übersicht, Pro Mitarbeiter, Pro Standort, Details."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl nicht installiert. Bitte 'pip install openpyxl' ausführen.",
        )

    now_berlin = datetime.now(_BERLIN)
    target_year = year or now_berlin.year
    target_month = month or now_berlin.month
    start_utc, end_utc = _month_bounds_utc(target_year, target_month)

    # ── Mitarbeiter-Filter ────────────────────────────────────────────────────
    emp_ids: set[int] | None = None
    if location_id is not None:
        emp_ids = _emp_ids_for_location(db, location_id)
    if employee_id is not None:
        emp_ids = (emp_ids & {employee_id}) if emp_ids is not None else {employee_id}

    emp_stmt = select(Employee).order_by(Employee.name)
    if emp_ids is not None:
        emp_stmt = emp_stmt.where(Employee.id.in_(list(emp_ids) or [-1]))
    employees_list = db.scalars(emp_stmt).all()

    # ── WorkSessions laden ────────────────────────────────────────────────────
    ws_stmt = (
        select(WorkSession)
        .where(WorkSession.checkin_time >= start_utc)
        .where(WorkSession.checkin_time < end_utc)
    )
    if emp_ids is not None:
        ws_stmt = ws_stmt.where(WorkSession.employee_id.in_(list(emp_ids) or [-1]))
    all_ws = db.scalars(ws_stmt).all()

    emp_name_map = {e.id: e.name for e in employees_list}
    emp_email_map = {e.id: e.email for e in employees_list}

    # ── Standort-Mapping ──────────────────────────────────────────────────────
    loc_map = {loc.id: loc.name for loc in db.scalars(select(WorkplaceLocation)).all()}
    emp_to_loc: dict[int, str] = {}
    for row in db.execute(
        select(EmployeeWorkLocation.employee_id, EmployeeWorkLocation.location_id)
        .order_by(EmployeeWorkLocation.location_id)
    ).all():
        if row.employee_id not in emp_to_loc:
            emp_to_loc[row.employee_id] = loc_map.get(row.location_id, "Unbekannt")
    for emp in employees_list:
        if emp.id not in emp_to_loc and emp.assigned_location_id:
            emp_to_loc[emp.id] = loc_map.get(emp.assigned_location_id, "Unbekannt")

    # ── openpyxl Styles ───────────────────────────────────────────────────────
    _HDR_FILL   = PatternFill("solid", fgColor="1E3A5F")
    _HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
    _GREEN_FILL = PatternFill("solid", fgColor="D1FAE5")
    _RED_FILL   = PatternFill("solid", fgColor="FEE2E2")
    _ORANGE_FILL= PatternFill("solid", fgColor="FEF3C7")
    _BLUE_FILL  = PatternFill("solid", fgColor="DBEAFE")
    _GRAY_FILL  = PatternFill("solid", fgColor="F1F5F9")

    def _header_row(ws_sheet, cols: list[str]) -> None:
        for ci, col in enumerate(cols, 1):
            cell = ws_sheet.cell(row=1, column=ci, value=col)
            cell.font = _HDR_FONT
            cell.fill = _HDR_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _autofit(ws_sheet) -> None:
        for col in ws_sheet.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws_sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    _STATUS_LABEL = {
        "approved":  "Genehmigt",
        "corrected": "Korrigiert",
        "rejected":  "Abgelehnt",
        "pending":   "Ausstehend",
    }
    _STATUS_FILL = {
        "approved":  _GREEN_FILL,
        "corrected": _BLUE_FILL,
        "rejected":  _RED_FILL,
        "pending":   _ORANGE_FILL,
    }

    wb = Workbook()

    # ── Sheet 1: Übersicht ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Übersicht"
    period_label = f"{target_month:02d}/{target_year}"
    official_secs = sum(ws.duration_seconds for ws in all_ws if ws.status in ("approved", "corrected"))
    pending_secs  = sum(ws.duration_seconds for ws in all_ws if ws.status == "pending")
    ws1.append(["KPI", "Wert"])
    _header_row(ws1, ["KPI", "Wert"])
    ws1.append(["Zeitraum", period_label])
    ws1.append(["Mitarbeiter", len(employees_list)])
    ws1.append(["Sitzungen gesamt", len(all_ws)])
    ws1.append(["Offizielle Gesamtstunden", round(official_secs / 3600, 2)])
    ws1.append(["Ausstehende Stunden", round(pending_secs / 3600, 2)])
    ws1.freeze_panes = "A2"
    _autofit(ws1)

    # ── Sheet 2: Pro Mitarbeiter ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Pro Mitarbeiter")
    cols2 = ["Mitarbeiter", "E-Mail", "Soll (h)", "Ist (h)", "Differenz (h)"]
    _header_row(ws2, cols2)
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(cols2))}1"

    total_target = 0.0
    total_official = 0.0
    for emp in employees_list:
        target_h = float(resolved_month_target_hours(emp))
        off_sec = sum(ws.duration_seconds for ws in all_ws if ws.employee_id == emp.id and ws.status in ("approved", "corrected"))
        off_h = round(off_sec / 3600, 2)
        diff = round(off_h - target_h, 2)
        row_idx = ws2.max_row + 1
        ws2.append([emp.name, emp.email, target_h, off_h, diff])
        fill = _GREEN_FILL if off_h >= target_h else _RED_FILL
        for ci in range(3, 6):
            ws2.cell(row=row_idx, column=ci).fill = fill
        total_target += target_h
        total_official += off_h

    # Summenzeile
    sum_row = ws2.max_row + 1
    ws2.append(["GESAMT", "", round(total_target, 2), round(total_official, 2), round(total_official - total_target, 2)])
    for ci in range(1, 6):
        ws2.cell(row=sum_row, column=ci).font = Font(bold=True)
    _autofit(ws2)

    # ── Sheet 3: Pro Standort ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Pro Standort")
    cols3 = ["Standort", "Offizielle Stunden", "Mitarbeiteranzahl"]
    _header_row(ws3, cols3)
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(cols3))}1"

    loc_agg: dict[str, dict] = {}
    for ws in all_ws:
        if ws.status not in ("approved", "corrected"):
            continue
        loc_name = emp_to_loc.get(ws.employee_id, "Kein Standort")
        if loc_name not in loc_agg:
            loc_agg[loc_name] = {"seconds": 0, "emp_ids": set()}
        loc_agg[loc_name]["seconds"] += ws.duration_seconds
        loc_agg[loc_name]["emp_ids"].add(ws.employee_id)

    for loc_name, data in sorted(loc_agg.items()):
        ws3.append([loc_name, round(data["seconds"] / 3600, 2), len(data["emp_ids"])])
    _autofit(ws3)

    # ── Sheet 4: Details ──────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Details")
    cols4 = ["Mitarbeiter", "E-Mail", "Standort", "Check-in", "Check-out", "Dauer (h)", "Status", "Genehmigung"]
    _header_row(ws4, cols4)
    ws4.freeze_panes = "A2"
    ws4.auto_filter.ref = f"A1:{get_column_letter(len(cols4))}1"

    for ws in sorted(all_ws, key=lambda x: x.checkin_time):
        checkin_str  = ws.checkin_time.astimezone(_BERLIN).strftime("%d.%m.%Y %H:%M") if ws.checkin_time else ""
        checkout_str = ws.checkout_time.astimezone(_BERLIN).strftime("%d.%m.%Y %H:%M") if ws.checkout_time else "offen"
        loc_name     = emp_to_loc.get(ws.employee_id, "—")
        row_idx = ws4.max_row + 1
        ws4.append([
            emp_name_map.get(ws.employee_id, f"ID {ws.employee_id}"),
            emp_email_map.get(ws.employee_id, ""),
            loc_name,
            checkin_str,
            checkout_str,
            round(ws.duration_seconds / 3600, 2),
            ws.status,
            _STATUS_LABEL.get(ws.status, ws.status),
        ])
        fill = _STATUS_FILL.get(ws.status, _GRAY_FILL)
        for ci in range(1, 9):
            ws4.cell(row=row_idx, column=ci).fill = fill
    _autofit(ws4)

    # ── Datei in Memory speichern und streamen ────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"time_stemple_{target_year}_{target_month:02d}_{now_str}.xlsx"

    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Report V2 – professionelles HR/Payroll-Reporting
# ═══════════════════════════════════════════════════════════════════════════════

def _build_v2_data(
    db: Session,
    employee_ids: list[int],
    location_ids: list[int],
    from_date: str,
    to_date: str,
    grouping: str,
) -> tuple[ReportV2Response, dict[int, str]]:
    """Returns (ReportV2Response, emp_email_map {id→email})."""
    try:
        from_d = date.fromisoformat(from_date)
        to_d   = date.fromisoformat(to_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Datumsformat ungültig (erwartet: YYYY-MM-DD).")
    if to_d < from_d:
        raise HTTPException(status_code=400, detail="to_date muss nach from_date liegen.")

    start_utc = datetime(from_d.year, from_d.month, from_d.day, tzinfo=_BERLIN).astimezone(UTC)
    end_utc   = (datetime(to_d.year, to_d.month, to_d.day, tzinfo=_BERLIN) + timedelta(days=1)).astimezone(UTC)

    all_emps: list[Employee] = db.scalars(
        select(Employee).where(Employee.is_active == True).order_by(Employee.name)
    ).all()
    emp_map: dict[int, Employee]  = {e.id: e for e in all_emps}
    emp_email_map: dict[int, str] = {e.id: e.email for e in all_emps}

    # Standort-Filter → zugehörige Mitarbeiter-IDs
    loc_emp_ids: set[int] | None = None
    if location_ids:
        loc_emp_ids = set()
        for lid in location_ids:
            loc_emp_ids |= _emp_ids_for_location(db, lid)

    if employee_ids and loc_emp_ids is not None:
        target_ids: set[int] = set(employee_ids) & loc_emp_ids
    elif employee_ids:
        target_ids = set(employee_ids)
    elif loc_emp_ids is not None:
        target_ids = loc_emp_ids
    else:
        target_ids = set(emp_map.keys())

    # Standort-Mapping
    loc_name_map: dict[int, str] = {
        loc.id: loc.name for loc in db.scalars(select(WorkplaceLocation)).all()
    }
    emp_to_loc: dict[int, str] = {}
    if location_ids:
        # Wenn Standort-Filter aktiv: Mitarbeiter dem gefilterten Standort zuordnen,
        # nicht ihrem "ersten" M2M-Eintrag (der wäre sonst immer die niedrigste location_id).
        for lid in location_ids:
            loc_name = loc_name_map.get(lid, "Unbekannt")
            for emp_id in _emp_ids_for_location(db, lid):
                if emp_id in target_ids and emp_id not in emp_to_loc:
                    emp_to_loc[emp_id] = loc_name
    else:
        for row in db.execute(
            select(EmployeeWorkLocation.employee_id, EmployeeWorkLocation.location_id)
            .order_by(EmployeeWorkLocation.location_id)
        ).all():
            if row.employee_id in target_ids and row.employee_id not in emp_to_loc:
                emp_to_loc[row.employee_id] = loc_name_map.get(row.location_id, "Unbekannt")
        for e in all_emps:
            if e.id in target_ids and e.id not in emp_to_loc and e.assigned_location_id:
                emp_to_loc[e.id] = loc_name_map.get(e.assigned_location_id, "Unbekannt")

    # WorkSessions laden
    safe_ids = list(target_ids) or [-1]
    ws_list: list[WorkSession] = db.scalars(
        select(WorkSession)
        .where(WorkSession.checkin_time >= start_utc)
        .where(WorkSession.checkin_time <  end_utc)
        .where(WorkSession.employee_id.in_(safe_ids))
        .order_by(WorkSession.checkin_time)
    ).all()

    # GPS-Fallback: Standort aus Check-in-Koordinaten ermitteln
    # (für Mitarbeiter ohne Standort-Zuweisung in employee_work_locations)
    all_locs_full = db.scalars(select(WorkplaceLocation)).all()
    checkin_log_ids = [ws.checkin_log_id for ws in ws_list if ws.checkin_log_id is not None]
    ws_gps_loc: dict[int, str] = {}
    if checkin_log_ids and all_locs_full:
        att_map = {
            a.id: a
            for a in db.scalars(
                select(Attendance).where(Attendance.id.in_(checkin_log_ids))
            ).all()
        }
        for ws in ws_list:
            if ws.checkin_log_id and ws.checkin_log_id in att_map:
                att = att_map[ws.checkin_log_id]
                for loc in all_locs_full:
                    if haversine_meters(att.lat, att.lng, loc.lat, loc.lng) <= float(loc.radius_meters):
                        ws_gps_loc[ws.id] = loc.name
                        break

    def _session_loc(ws: WorkSession) -> str:
        # GPS-Treffer zuerst (wo war der Mitarbeiter bei DIESEM Check-in?)
        # Zuweisung nur als Fallback wenn kein GPS-Match vorhanden
        return ws_gps_loc.get(ws.id) or emp_to_loc.get(ws.employee_id) or "Kein Standort"

    # Session-Rows
    session_rows: list[ReportV2SessionRow] = []
    for ws in ws_list:
        emp = emp_map.get(ws.employee_id)
        if emp is None:
            continue
        ci_b    = ws.checkin_time.astimezone(_BERLIN)
        dur_min = ws.duration_seconds // 60
        session_rows.append(ReportV2SessionRow(
            employee_id=ws.employee_id,
            employee_name=emp.name,
            date=ci_b.date().isoformat(),
            weekday=_WEEKDAYS_DE[ci_b.weekday()],
            location_name=_session_loc(ws),
            checkin_time=ws.checkin_time,
            checkout_time=ws.checkout_time,
            break_minutes=0,
            work_minutes=dur_min,
            duration_minutes=dur_min,
            status=ws.status,
        ))

    # KPIs — aggregate using raw duration_seconds to avoid minute-truncation errors
    ws_by_id: dict[int, WorkSession] = {ws.employee_id: ws for ws in ws_list}
    off_sec  = sum(ws.duration_seconds for ws in ws_list if ws.status in ("approved", "corrected"))
    pend_sec = sum(ws.duration_seconds for ws in ws_list if ws.status == "pending")
    tot_sec  = sum(ws.duration_seconds for ws in ws_list)
    date_set = {r.date for r in session_rows}

    kpis = ReportV2KPIs(
        total_hours=round(tot_sec / 3600, 2),
        official_hours=round(off_sec / 3600, 2),
        pending_hours=round(pend_sec / 3600, 2),
        total_shifts=len(session_rows),
        location_count=len({r.location_name for r in session_rows}),
        work_days=len(date_set),
    )

    # Standort-Zusammenfassung — nur approved/corrected Sessions zählen
    loc_agg: dict[str, dict] = {}
    for ws in ws_list:
        if ws.status not in ("approved", "corrected"):
            continue
        loc_name = _session_loc(ws)
        if loc_name not in loc_agg:
            loc_agg[loc_name] = {"cnt": 0, "sec": 0}
        loc_agg[loc_name]["cnt"] += 1
        loc_agg[loc_name]["sec"] += ws.duration_seconds
    location_summary = sorted([
        ReportV2LocationRow(
            location_name=k,
            shift_count=v["cnt"],
            total_hours=round(v["sec"] / 3600, 2),
        ) for k, v in loc_agg.items()
    ], key=lambda x: x.total_hours, reverse=True)

    # Trend-Daten — use seconds from ws_list
    trend_agg: dict[str, dict] = {}
    for ws in ws_list:
        ci_b = ws.checkin_time.astimezone(_BERLIN)
        if grouping == "daily":
            key   = ci_b.strftime("%Y-%m-%d")
            label = ci_b.strftime("%d.%m.%Y")
        elif grouping == "weekly":
            iso   = ci_b.isocalendar()
            key   = f"{iso.year}-W{iso.week:02d}"
            label = f"KW {iso.week:02d}/{str(iso.year)[2:]}"
        else:
            key   = ci_b.strftime("%Y-%m")
            label = ci_b.strftime("%m/%Y")
        if key not in trend_agg:
            trend_agg[key] = {"label": label, "off": 0, "pend": 0}
        if ws.status in ("approved", "corrected"):
            trend_agg[key]["off"]  += ws.duration_seconds
        elif ws.status == "pending":
            trend_agg[key]["pend"] += ws.duration_seconds
    trend_data = [
        ReportV2TrendRow(
            period=k, period_label=v["label"],
            official_hours=round(v["off"] / 3600, 2),
            pending_hours=round(v["pend"] / 3600, 2),
        ) for k, v in sorted(trend_agg.items())
    ]

    # Mitarbeiter-Zusammenfassung — use seconds from ws_list
    emp_agg: dict[int, dict] = {}
    for ws in ws_list:
        emp = emp_map.get(ws.employee_id)
        if emp is None:
            continue
        if ws.employee_id not in emp_agg:
            emp_agg[ws.employee_id] = {"name": emp.name, "off": 0, "pend": 0, "shifts": 0, "dates": set()}
        if ws.status in ("approved", "corrected"):
            emp_agg[ws.employee_id]["off"]  += ws.duration_seconds
        elif ws.status == "pending":
            emp_agg[ws.employee_id]["pend"] += ws.duration_seconds
        emp_agg[ws.employee_id]["shifts"] += 1
        emp_agg[ws.employee_id]["dates"].add(ws.checkin_time.astimezone(_BERLIN).date().isoformat())
    employee_summary = sorted([
        ReportV2EmployeeRow(
            employee_id=eid,
            employee_name=v["name"],
            official_hours=round(v["off"] / 3600, 2),
            pending_hours=round(v["pend"] / 3600, 2),
            target_hours=resolved_month_target_hours(emp_map[eid]) if eid in emp_map else 160,
            diff_hours=round(
                v["off"] / 3600 - (resolved_month_target_hours(emp_map[eid]) if eid in emp_map else 160), 2
            ),
            shift_count=v["shifts"],
            work_days=len(v["dates"]),
        ) for eid, v in emp_agg.items()
    ], key=lambda x: x.employee_name)

    total_target = sum(
        resolved_month_target_hours(emp_map[eid]) for eid in target_ids if eid in emp_map
    )
    period_summary = ReportV2PeriodSummary(
        total_hours=round(tot_sec / 3600, 2),
        official_hours=round(off_sec / 3600, 2),
        pending_hours=round(pend_sec / 3600, 2),
        target_hours=total_target,
        diff_hours=round(off_sec / 3600 - total_target, 2),
        shift_count=len(session_rows),
        work_days=len(date_set),
    )

    return (
        ReportV2Response(
            kpis=kpis,
            sessions=session_rows,
            location_summary=location_summary,
            period_summary=period_summary,
            trend_data=trend_data,
            employee_summary=employee_summary,
        ),
        emp_email_map,
    )


@router.get("/v2/summary", response_model=ReportV2Response)
def reports_v2_summary(
    employee_ids: list[int] = Query(default=[]),
    location_ids: list[int] = Query(default=[]),
    from_date: str = Query(..., description="YYYY-MM-DD"),
    to_date:   str = Query(..., description="YYYY-MM-DD"),
    grouping:  str = Query(default="daily"),
    db: Session = Depends(get_db),
    _: Employee = Depends(require_admin),
):
    """Professioneller V2-Arbeitszeitbericht: Multi-Filter, Trend, Standort- und Mitarbeiterauswertung."""
    data, _ = _build_v2_data(db, employee_ids, location_ids, from_date, to_date, grouping)
    return data


@router.get("/v2/excel")
def reports_v2_excel(
    employee_ids: list[int] = Query(default=[]),
    location_ids: list[int] = Query(default=[]),
    from_date: str = Query(..., description="YYYY-MM-DD"),
    to_date:   str = Query(..., description="YYYY-MM-DD"),
    grouping:  str = Query(default="daily"),
    db: Session = Depends(get_db),
    _: Employee = Depends(require_admin),
):
    """Excel-Export V2 — professionelles Design mit Charts, Deckblatt und Farbschema."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment, Border, Font, PatternFill, Side,
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference
        from openpyxl.chart.series import DataPoint
    except ImportError as exc:
        logger.error("openpyxl not available: %s", exc)
        raise HTTPException(status_code=500, detail="openpyxl nicht installiert.")

    try:
        report, emp_email_map = _build_v2_data(
            db, employee_ids, location_ids, from_date, to_date, grouping
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("_build_v2_data failed:\n%s", traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Datenabruf fehlgeschlagen: {exc}")

    # ═══════════════════════════════════════════════════════
    # HELPERS
    # ═══════════════════════════════════════════════════════
    def _to_berlin(dt: datetime | None) -> datetime | None:
        if dt is None:
            return None
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=UTC)
        return dt.astimezone(_BERLIN)

    def _h(secs: float) -> float:
        """Stunden als gerundeter Float — für Zahlenfelder."""
        return round(secs / 3600, 2) if secs else 0.0

    def _fmt_h(h: float) -> str:
        """'55,43 h' für Anzeige-Felder."""
        return f"{h:,.2f} h".replace(",", "X").replace(".", ",").replace("X", ".")

    def _fmt_hm(minutes: float | int | None) -> str:
        """'55:43' Format."""
        if minutes is None:
            return "—"
        minutes = int(round(minutes))
        sign = "-" if minutes < 0 else ""
        hh, mm = divmod(abs(minutes), 60)
        return f"{sign}{hh}:{mm:02d}"

    def _autofit(ws_sheet, min_w: int = 8, max_w: int = 42) -> None:
        try:
            for col in ws_sheet.columns:
                if not col:
                    continue
                w = max((len(str(cell.value or "")) for cell in col), default=min_w)
                ws_sheet.column_dimensions[
                    get_column_letter(col[0].column)
                ].width = min(max(w + 3, min_w), max_w)
        except Exception:
            pass

    def _thin_border() -> Border:
        s = Side(style="thin", color="FFD1D9E6")
        return Border(left=s, right=s, top=s, bottom=s)

    def _outer_border() -> Border:
        s = Side(style="medium", color="FF1E3A5F")
        return Border(left=s, right=s, top=s, bottom=s)

    def _apply_table_border(ws_sheet, row_idx: int, num_cols: int) -> None:
        b = _thin_border()
        for ci in range(1, num_cols + 1):
            ws_sheet.cell(row=row_idx, column=ci).border = b

    try:
        # ═══════════════════════════════════════════════════════
        # FARB-PALETTE & STYLES
        # ═══════════════════════════════════════════════════════
        C_NAVY       = "FF1E3A5F"
        C_NAVY_LIGHT = "FF2D5282"
        C_GOLD       = "FFC8A84B"
        C_WHITE      = "FFFFFFFF"
        C_ROW_ALT    = "FFF0F4FA"
        C_ROW_WHITE  = "FFFFFFFF"
        C_GREEN_D    = "FF065F46"
        C_GREEN_BG   = "FFD1FAE5"
        C_RED_D      = "FF991B1B"
        C_RED_BG     = "FFFEE2E2"
        C_ORANGE_BG  = "FFFEF3C7"
        C_BLUE_BG    = "FFDBEAFE"
        C_GRAY_BG    = "FFF8FAFC"
        C_SECTION    = "FFE8EDF5"

        def _fill(hex_color: str) -> PatternFill:
            return PatternFill("solid", fgColor=hex_color)

        def _font(bold=False, size=10, color=C_NAVY, italic=False) -> Font:
            return Font(bold=bold, size=size, color=color, italic=italic,
                        name="Calibri")

        HDR_FILL  = _fill(C_NAVY)
        HDR_FONT  = _font(bold=True, size=10, color=C_WHITE)
        HDR_ALN   = Alignment(horizontal="center", vertical="center", wrap_text=True)

        SUBHDR_FILL = _fill(C_NAVY_LIGHT)
        SUBHDR_FONT = _font(bold=True, size=9, color=C_WHITE)

        BOLD10 = _font(bold=True, size=10)
        BOLD9  = _font(bold=True, size=9)
        REG9   = _font(size=9)
        MUTED  = _font(size=9, color="FF64748B")

        STATUS_FILL = {
            "approved":  _fill(C_GREEN_BG),
            "corrected": _fill(C_BLUE_BG),
            "rejected":  _fill(C_RED_BG),
            "pending":   _fill(C_ORANGE_BG),
        }
        STATUS_FONT = {
            "approved":  _font(bold=True, size=9, color="FF065F46"),
            "corrected": _font(bold=True, size=9, color="FF1E40AF"),
            "rejected":  _font(bold=True, size=9, color=C_RED_D),
            "pending":   _font(bold=True, size=9, color="FF92400E"),
        }
        STATUS_DE = {
            "approved":  "✔ Genehmigt",
            "corrected": "✎ Korrigiert",
            "rejected":  "✖ Abgelehnt",
            "pending":   "⏳ Ausstehend",
        }

        ps        = report.period_summary
        k         = report.kpis
        multi_emp = len(report.employee_summary) > 1
        now_str   = datetime.now(_BERLIN).strftime("%d.%m.%Y %H:%M")

        wb = Workbook()

        # ═══════════════════════════════════════════════════════
        # SHEET 1 — DECKBLATT
        # ═══════════════════════════════════════════════════════
        ws1       = wb.active
        ws1.title = "📊 Übersicht"
        ws1.sheet_view.showGridLines = False
        ws1.column_dimensions["A"].width = 3
        ws1.column_dimensions["B"].width = 28
        ws1.column_dimensions["C"].width = 22
        ws1.column_dimensions["D"].width = 22
        ws1.column_dimensions["E"].width = 22
        ws1.column_dimensions["F"].width = 22
        ws1.column_dimensions["G"].width = 3

        # Titel-Banner (Zeilen 1-4)
        for r in range(1, 5):
            ws1.row_dimensions[r].height = 18
        for c in range(1, 8):
            for r in range(1, 5):
                ws1.cell(row=r, column=c).fill = _fill(C_NAVY)

        ws1.merge_cells("B1:F4")
        title_cell = ws1["B1"]
        title_cell.value     = "ARBEITSZEITBERICHT"
        title_cell.font      = Font(bold=True, size=22, color=C_WHITE, name="Calibri")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")

        # Gold-Linie
        ws1.row_dimensions[5].height = 4
        for c in range(1, 8):
            ws1.cell(row=5, column=c).fill = _fill(C_GOLD)

        # Meta-Info (Zeilen 6-9)
        ws1.row_dimensions[6].height = 6
        meta = [
            (7,  "Zeitraum",    f"{from_date}  →  {to_date}"),
            (8,  "Erstellt am", now_str),
            (9,  "Mitarbeiter", str(len(report.employee_summary)) if report.employee_summary else "Alle"),
        ]
        for row_n, label, val in meta:
            ws1.row_dimensions[row_n].height = 20
            lc = ws1.cell(row=row_n, column=2, value=label)
            lc.font      = _font(bold=True, size=10, color="FF64748B")
            lc.alignment = Alignment(vertical="center")
            vc = ws1.cell(row=row_n, column=3, value=val)
            vc.font      = _font(bold=True, size=11, color=C_NAVY)
            vc.alignment = Alignment(vertical="center")

        # KPI-Karten (Zeilen 11-17)
        ws1.row_dimensions[10].height = 10
        ws1.row_dimensions[11].height = 14
        ws1.row_dimensions[12].height = 36
        ws1.row_dimensions[13].height = 18
        ws1.row_dimensions[14].height = 10
        ws1.row_dimensions[15].height = 14
        ws1.row_dimensions[16].height = 36
        ws1.row_dimensions[17].height = 18

        kpi_cards = [
            ("B",  "C",  "GESAMTSTUNDEN",           _fmt_h(k.total_hours),          C_NAVY),
            ("D",  "E",  "OFFIZIELLE STUNDEN",       _fmt_h(ps.official_hours),      "FF065F46"),
            ("F",  "G",  "AUSSTEHEND",               _fmt_h(ps.pending_hours),       "FF92400E"),
        ]
        kpi_cards2 = [
            ("B",  "C",  "SOLL-STUNDEN",             f"{ps.target_hours} h",         C_NAVY_LIGHT),
            ("D",  "E",  "DIFFERENZ",                _fmt_hm(ps.diff_hours * 60),    C_GREEN_D if ps.diff_hours >= 0 else C_RED_D),
            ("F",  "G",  "SCHICHTEN",                str(k.total_shifts),             "FF6B21A8"),
        ]

        for start_col, end_col, label, val, color in kpi_cards:
            ws1.merge_cells(f"{start_col}11:{end_col}11")
            lbl = ws1[f"{start_col}11"]
            lbl.value     = label
            lbl.font      = _font(bold=True, size=8, color="FF64748B")
            lbl.alignment = Alignment(horizontal="center", vertical="center")
            lbl.fill      = _fill(C_SECTION)

            ws1.merge_cells(f"{start_col}12:{end_col}12")
            vc = ws1[f"{start_col}12"]
            vc.value     = val
            vc.font      = Font(bold=True, size=18, color=color, name="Calibri")
            vc.alignment = Alignment(horizontal="center", vertical="center")
            vc.fill      = _fill(C_SECTION)

            ws1.merge_cells(f"{start_col}13:{end_col}13")
            ws1[f"{start_col}13"].fill = _fill(C_SECTION)

            # farbiger Balken unter KPI
            ws1.merge_cells(f"{start_col}14:{end_col}14")
            ws1[f"{start_col}14"].fill = _fill(color)

        for start_col, end_col, label, val, color in kpi_cards2:
            ws1.merge_cells(f"{start_col}15:{end_col}15")
            lbl = ws1[f"{start_col}15"]
            lbl.value     = label
            lbl.font      = _font(bold=True, size=8, color="FF64748B")
            lbl.alignment = Alignment(horizontal="center", vertical="center")
            lbl.fill      = _fill(C_SECTION)

            ws1.merge_cells(f"{start_col}16:{end_col}16")
            vc = ws1[f"{start_col}16"]
            vc.value     = val
            vc.font      = Font(bold=True, size=18, color=color, name="Calibri")
            vc.alignment = Alignment(horizontal="center", vertical="center")
            vc.fill      = _fill(C_SECTION)

            ws1.merge_cells(f"{start_col}17:{end_col}17")
            ws1[f"{start_col}17"].fill = _fill(C_SECTION)

            ws1.merge_cells(f"{start_col}18:{end_col}18")
            ws1[f"{start_col}18"].fill = _fill(color)

        # Weitere Kennzahlen (Zeilen 20-26)
        ws1.row_dimensions[19].height = 10
        ws1.row_dimensions[20].height = 20
        stat_hdr = ws1.cell(row=20, column=2, value="WEITERE KENNZAHLEN")
        stat_hdr.font      = _font(bold=True, size=9, color="FF64748B")
        stat_hdr.alignment = Alignment(vertical="center")

        stat_rows = [
            ("Arbeitstage",      str(k.work_days)),
            ("Standorte",        str(k.location_count)),
            ("Zeitraum (Tage)",  str((date.fromisoformat(to_date) - date.fromisoformat(from_date)).days + 1)),
        ]
        for i, (lbl, val) in enumerate(stat_rows, start=21):
            ws1.row_dimensions[i].height = 18
            lc = ws1.cell(row=i, column=2, value=lbl)
            lc.font      = REG9
            lc.alignment = Alignment(vertical="center")
            vc = ws1.cell(row=i, column=3, value=val)
            vc.font      = BOLD9
            vc.alignment = Alignment(vertical="center")
            for c in range(2, 7):
                ws1.cell(row=i, column=c).fill = _fill(C_ROW_ALT) if i % 2 == 0 else _fill(C_ROW_WHITE)

        # Footer
        ws1.row_dimensions[27].height = 6
        for c in range(1, 8):
            ws1.cell(row=27, column=c).fill = _fill(C_GOLD)
        ws1.row_dimensions[28].height = 16
        ft = ws1.cell(row=28, column=2, value="Erstellt mit Time Stemple Workforce Management")
        ft.font      = _font(italic=True, size=8, color="FF94A3B8")
        ft.alignment = Alignment(vertical="center")
        ws1.merge_cells("B28:F28")

        # ═══════════════════════════════════════════════════════
        # SHEET 2 — SCHICHTDETAILS
        # ═══════════════════════════════════════════════════════
        ws2       = wb.create_sheet("📋 Schichtdetails")
        ws2.sheet_view.showGridLines = False
        ws2.freeze_panes = "A3"

        # Titel-Zeile
        ws2.row_dimensions[1].height = 28
        col_offset = 2 if multi_emp else 0
        total_cols = col_offset + 8
        ws2.merge_cells(f"A1:{get_column_letter(total_cols)}1")
        th = ws2["A1"]
        th.value     = "SCHICHTDETAILS"
        th.font      = Font(bold=True, size=13, color=C_WHITE, name="Calibri")
        th.fill      = _fill(C_NAVY)
        th.alignment = Alignment(horizontal="left", vertical="center",
                                  indent=1)

        # Spalten-Header (Zeile 2)
        ws2.row_dimensions[2].height = 24
        base_cols = ["Datum", "Wochentag", "Standort", "Check-In", "Check-Out",
                     "Pause", "Std. (h)", "Status"]
        cols2 = (["Mitarbeiter", "E-Mail"] if multi_emp else []) + base_cols
        for ci, c in enumerate(cols2, 1):
            cell = ws2.cell(row=2, column=ci, value=c)
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = HDR_ALN
            cell.border    = _thin_border()

        ws2.auto_filter.ref = f"A2:{get_column_letter(len(cols2))}2"

        # Daten
        for idx, row in enumerate(report.sessions):
            ci_b = _to_berlin(row.checkin_time)
            co_b = _to_berlin(row.checkout_time)
            if ci_b is None:
                continue
            ri   = ws2.max_row + 1
            ws2.row_dimensions[ri].height = 18
            alt  = _fill(C_ROW_ALT) if ri % 2 == 0 else _fill(C_ROW_WHITE)
            dur_h = round(row.work_minutes / 60, 2) if row.work_minutes else 0.0
            vals = (
                ([row.employee_name, emp_email_map.get(row.employee_id, "")] if multi_emp else [])
                + [
                    ci_b.strftime("%d.%m.%Y"),
                    row.weekday,
                    row.location_name,
                    ci_b.strftime("%H:%M"),
                    co_b.strftime("%H:%M") if co_b else "—",
                    "—",
                    dur_h,
                    STATUS_DE.get(row.status, row.status),
                ]
            )
            for ci, v in enumerate(vals, 1):
                cell = ws2.cell(row=ri, column=ci, value=v)
                cell.border = _thin_border()
                # Status-Spalte einfärben
                if ci == len(vals):
                    cell.fill = STATUS_FILL.get(row.status, _fill(C_GRAY_BG))
                    cell.font = STATUS_FONT.get(row.status, REG9)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif ci <= col_offset:
                    cell.font = BOLD9
                    cell.fill = alt
                    cell.alignment = Alignment(vertical="center")
                else:
                    cell.fill = alt
                    cell.font = REG9
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Stunden numerisch formatieren
            h_col = len(vals) - 1
            ws2.cell(row=ri, column=h_col).number_format = '#,##0.00" h"'

        # Summenzeile
        if ws2.max_row >= 3:
            sr = ws2.max_row + 1
            ws2.row_dimensions[sr].height = 20
            for ci in range(1, len(cols2) + 1):
                cell = ws2.cell(row=sr, column=ci)
                cell.fill   = _fill(C_NAVY)
                cell.border = _thin_border()
            h_col   = len(cols2) - 1
            sum_val = sum(
                (ws2.cell(row=r, column=h_col).value or 0)
                for r in range(3, sr)
                if isinstance(ws2.cell(row=r, column=h_col).value, (int, float))
            )
            lc = ws2.cell(row=sr, column=1, value="GESAMT")
            lc.font      = HDR_FONT
            lc.alignment = Alignment(vertical="center", indent=1)
            sc = ws2.cell(row=sr, column=h_col, value=round(sum_val, 2))
            sc.font          = HDR_FONT
            sc.alignment     = Alignment(horizontal="center", vertical="center")
            sc.number_format = '#,##0.00" h"'

        _autofit(ws2)
        ws2.column_dimensions["A"].width = max(ws2.column_dimensions["A"].width, 12)

        # ═══════════════════════════════════════════════════════
        # SHEET 3 — MITARBEITERÜBERSICHT + BALKENDIAGRAMM
        # ═══════════════════════════════════════════════════════
        ws3       = wb.create_sheet("👥 Mitarbeiter")
        ws3.sheet_view.showGridLines = False
        ws3.freeze_panes = "A3"

        # Titel
        ws3.row_dimensions[1].height = 28
        ws3.merge_cells("A1:H1")
        th3 = ws3["A1"]
        th3.value     = "MITARBEITERÜBERSICHT"
        th3.font      = Font(bold=True, size=13, color=C_WHITE, name="Calibri")
        th3.fill      = _fill(C_NAVY)
        th3.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Header
        ws3.row_dimensions[2].height = 24
        cols3 = ["Mitarbeiter", "E-Mail", "Ist (h)", "Ausstehend (h)",
                 "Soll (h)", "Differenz (h)", "Schichten", "Arbeitstage"]
        for ci, c in enumerate(cols3, 1):
            cell = ws3.cell(row=2, column=ci, value=c)
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = HDR_ALN
            cell.border    = _thin_border()

        ws3.auto_filter.ref = f"A2:{get_column_letter(len(cols3))}2"

        data_start3 = 3
        for idx, emp_row in enumerate(report.employee_summary):
            ri   = ws3.max_row + 1
            diff = emp_row.diff_hours or 0.0
            ws3.row_dimensions[ri].height = 18
            alt  = _fill(C_ROW_ALT) if ri % 2 == 0 else _fill(C_ROW_WHITE)
            vals = [
                emp_row.employee_name,
                emp_email_map.get(emp_row.employee_id, ""),
                round(emp_row.official_hours, 2),
                round(emp_row.pending_hours, 2),
                float(emp_row.target_hours),
                round(diff, 2),
                emp_row.shift_count,
                emp_row.work_days,
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws3.cell(row=ri, column=ci, value=v)
                cell.border = _thin_border()
                cell.font   = BOLD9 if ci == 1 else REG9
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="center" if ci > 2 else "left",
                )
                if ci == 1:
                    cell.fill = alt
                elif ci == 3:
                    cell.fill = _fill(C_GREEN_BG)
                    cell.font = _font(bold=True, size=9, color=C_GREEN_D)
                    cell.number_format = '#,##0.00" h"'
                elif ci == 4:
                    cell.fill = _fill(C_ORANGE_BG)
                    cell.number_format = '#,##0.00" h"'
                elif ci == 5:
                    cell.fill = _fill(C_SECTION)
                    cell.number_format = '#,##0.00" h"'
                elif ci == 6:
                    cell.fill = _fill(C_GREEN_BG) if diff >= 0 else _fill(C_RED_BG)
                    cell.font = _font(bold=True, size=9,
                                      color=C_GREEN_D if diff >= 0 else C_RED_D)
                    cell.number_format = '+#,##0.00" h";-#,##0.00" h"'
                else:
                    cell.fill = alt

        # Summenzeile
        if ws3.max_row >= 3:
            sr3 = ws3.max_row + 1
            ws3.row_dimensions[sr3].height = 20
            for ci in range(1, len(cols3) + 1):
                c = ws3.cell(row=sr3, column=ci)
                c.fill   = _fill(C_NAVY)
                c.border = _thin_border()
                c.font   = HDR_FONT
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws3.cell(row=sr3, column=1, value="GESAMT").alignment = Alignment(
                horizontal="left", vertical="center", indent=1)
            for ci, col_key in [(3, "official_hours"), (4, "pending_hours"),
                                 (5, "target_hours"), (7, "shift_count"), (8, "work_days")]:
                total = sum(getattr(e, col_key, 0) or 0 for e in report.employee_summary)
                c = ws3.cell(row=sr3, column=ci, value=round(total, 2))
                if ci in (3, 4, 5):
                    c.number_format = '#,##0.00" h"'
            diff_total = round(sum((e.diff_hours or 0) for e in report.employee_summary), 2)
            dc = ws3.cell(row=sr3, column=6, value=diff_total)
            dc.number_format = '+#,##0.00" h";-#,##0.00" h"'

        _autofit(ws3)

        # Balkendiagramm: Ist vs. Soll pro Mitarbeiter
        if report.employee_summary:
            data_rows3 = ws3.max_row - len(report.employee_summary)
            chart3 = BarChart()
            chart3.type    = "col"
            chart3.title   = "Ist- vs. Soll-Stunden pro Mitarbeiter"
            chart3.y_axis.title = "Stunden"
            chart3.x_axis.title = "Mitarbeiter"
            chart3.style   = 10
            chart3.width   = 22
            chart3.height  = 14

            n_emp = len(report.employee_summary)
            ist_ref  = Reference(ws3, min_col=3, min_row=2,
                                  max_row=2 + n_emp)
            soll_ref = Reference(ws3, min_col=5, min_row=2,
                                  max_row=2 + n_emp)
            cats_ref = Reference(ws3, min_col=1, min_row=3,
                                  max_row=2 + n_emp)
            chart3.add_data(ist_ref,  titles_from_data=True)
            chart3.add_data(soll_ref, titles_from_data=True)
            chart3.set_categories(cats_ref)
            chart3.series[0].graphicalProperties.solidFill = "1E3A5F"
            chart3.series[1].graphicalProperties.solidFill = "C8A84B"
            chart3.shape = 4
            ws3.add_chart(chart3, f"A{ws3.max_row + 2}")

        # ═══════════════════════════════════════════════════════
        # SHEET 4 — STANDORTAUSWERTUNG + DIAGRAMME
        # ═══════════════════════════════════════════════════════
        ws4       = wb.create_sheet("📍 Standorte")
        ws4.sheet_view.showGridLines = False
        ws4.freeze_panes = "A3"

        ws4.row_dimensions[1].height = 28
        ws4.merge_cells("A1:E1")
        th4 = ws4["A1"]
        th4.value     = "STANDORTAUSWERTUNG"
        th4.font      = Font(bold=True, size=13, color=C_WHITE, name="Calibri")
        th4.fill      = _fill(C_NAVY)
        th4.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws4.row_dimensions[2].height = 24
        cols4 = ["Standort", "Schichten", "Offizielle Std. (h)", "Anteil (%)", "Bewertung"]
        for ci, c in enumerate(cols4, 1):
            cell = ws4.cell(row=2, column=ci, value=c)
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = HDR_ALN
            cell.border    = _thin_border()

        total_h4 = sum(l.total_hours for l in report.location_summary) or 1.0
        RATINGS = ["★★★★★", "★★★★☆", "★★★☆☆", "★★☆☆☆", "★☆☆☆☆"]
        for idx, loc in enumerate(report.location_summary):
            ri  = ws4.max_row + 1
            ws4.row_dimensions[ri].height = 18
            alt = _fill(C_ROW_ALT) if ri % 2 == 0 else _fill(C_ROW_WHITE)
            pct = round(loc.total_hours / total_h4 * 100, 1)
            rating_idx = min(int((1 - pct / 100) * 5), 4)
            vals = [loc.location_name, loc.shift_count,
                    round(loc.total_hours, 2), pct, RATINGS[rating_idx]]
            for ci, v in enumerate(vals, 1):
                cell = ws4.cell(row=ri, column=ci, value=v)
                cell.border    = _thin_border()
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="center" if ci > 1 else "left",
                )
                if ci == 1:
                    cell.font = BOLD9
                    cell.fill = alt
                elif ci == 3:
                    cell.fill = _fill(C_GREEN_BG)
                    cell.font = _font(bold=True, size=9, color=C_GREEN_D)
                    cell.number_format = '#,##0.00" h"'
                elif ci == 4:
                    cell.fill = alt
                    cell.font = REG9
                    cell.number_format = '0.0"%"'
                elif ci == 5:
                    cell.fill = _fill(C_GOLD)
                    cell.font = Font(size=10, name="Calibri", color=C_NAVY)
                else:
                    cell.fill = alt
                    cell.font = REG9

        # Summenzeile
        if ws4.max_row >= 3:
            sr4 = ws4.max_row + 1
            ws4.row_dimensions[sr4].height = 20
            for ci in range(1, len(cols4) + 1):
                c = ws4.cell(row=sr4, column=ci)
                c.fill   = _fill(C_NAVY)
                c.border = _thin_border()
                c.font   = HDR_FONT
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws4.cell(row=sr4, column=1, value="GESAMT").alignment = Alignment(
                horizontal="left", vertical="center", indent=1)
            ws4.cell(row=sr4, column=2,
                     value=sum(l.shift_count for l in report.location_summary))
            th_c = ws4.cell(row=sr4, column=3,
                            value=round(sum(l.total_hours for l in report.location_summary), 2))
            th_c.number_format = '#,##0.00" h"'
            ws4.cell(row=sr4, column=4, value=100.0).number_format = '0.0"%"'

        _autofit(ws4)

        # Balkendiagramm Stunden pro Standort
        if report.location_summary:
            n_loc   = len(report.location_summary)
            bar4    = BarChart()
            bar4.type   = "col"
            bar4.title  = "Offizielle Stunden pro Standort"
            bar4.y_axis.title = "Stunden"
            bar4.style  = 10
            bar4.width  = 18
            bar4.height = 12
            h_ref4 = Reference(ws4, min_col=3, min_row=2, max_row=2 + n_loc)
            c_ref4 = Reference(ws4, min_col=1, min_row=3, max_row=2 + n_loc)
            bar4.add_data(h_ref4, titles_from_data=True)
            bar4.set_categories(c_ref4)
            bar4.series[0].graphicalProperties.solidFill = "1E3A5F"
            chart_row4 = ws4.max_row + 2
            ws4.add_chart(bar4, f"A{chart_row4}")

            # Tortendiagramm rechts daneben
            pie4        = PieChart()
            pie4.title  = "Verteilung nach Standort"
            pie4.style  = 10
            pie4.width  = 14
            pie4.height = 12
            p_ref4 = Reference(ws4, min_col=3, min_row=2, max_row=2 + n_loc)
            l_ref4 = Reference(ws4, min_col=1, min_row=3, max_row=2 + n_loc)
            pie4.add_data(p_ref4, titles_from_data=True)
            pie4.set_categories(l_ref4)
            PIE_COLORS = ["1E3A5F", "C8A84B", "0D9488", "7C3AED", "DC2626",
                          "2563EB", "D97706", "059669"]
            for i in range(min(n_loc, len(PIE_COLORS))):
                pt = DataPoint(idx=i)
                pt.graphicalProperties.solidFill = PIE_COLORS[i % len(PIE_COLORS)]
                pie4.series[0].dPt.append(pt)
            ws4.add_chart(pie4, f"J{chart_row4}")

        # ═══════════════════════════════════════════════════════
        # SHEET 5 — TRENDANALYSE + LINIENDIAGRAMM
        # ═══════════════════════════════════════════════════════
        ws5       = wb.create_sheet("📈 Trend")
        ws5.sheet_view.showGridLines = False
        ws5.freeze_panes = "A3"

        ws5.row_dimensions[1].height = 28
        ws5.merge_cells("A1:E1")
        th5 = ws5["A1"]
        th5.value     = "TRENDANALYSE"
        th5.font      = Font(bold=True, size=13, color=C_WHITE, name="Calibri")
        th5.fill      = _fill(C_NAVY)
        th5.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws5.row_dimensions[2].height = 24
        cols5 = ["Periode", "Offizielle Std. (h)", "Ausstehend (h)", "Gesamt (h)"]
        for ci, c in enumerate(cols5, 1):
            cell = ws5.cell(row=2, column=ci, value=c)
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = HDR_ALN
            cell.border    = _thin_border()

        for idx, t in enumerate(report.trend_data):
            ri  = ws5.max_row + 1
            ws5.row_dimensions[ri].height = 18
            alt = _fill(C_ROW_ALT) if ri % 2 == 0 else _fill(C_ROW_WHITE)
            tot = round((t.official_hours or 0) + (t.pending_hours or 0), 2)
            vals = [t.period_label,
                    round(t.official_hours or 0, 2),
                    round(t.pending_hours or 0, 2),
                    tot]
            for ci, v in enumerate(vals, 1):
                cell = ws5.cell(row=ri, column=ci, value=v)
                cell.border    = _thin_border()
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="center" if ci > 1 else "left",
                )
                if ci == 1:
                    cell.font = BOLD9
                    cell.fill = alt
                elif ci == 2:
                    cell.fill = _fill(C_GREEN_BG)
                    cell.font = _font(size=9, color=C_GREEN_D)
                    cell.number_format = '#,##0.00" h"'
                elif ci == 3:
                    cell.fill = _fill(C_ORANGE_BG)
                    cell.font = REG9
                    cell.number_format = '#,##0.00" h"'
                else:
                    cell.fill = alt
                    cell.font = REG9
                    cell.number_format = '#,##0.00" h"'

        # Summenzeile
        if ws5.max_row >= 3:
            sr5 = ws5.max_row + 1
            ws5.row_dimensions[sr5].height = 20
            for ci in range(1, len(cols5) + 1):
                c = ws5.cell(row=sr5, column=ci)
                c.fill   = _fill(C_NAVY)
                c.border = _thin_border()
                c.font   = HDR_FONT
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws5.cell(row=sr5, column=1, value="GESAMT").alignment = Alignment(
                horizontal="left", vertical="center", indent=1)
            for ci in [2, 3, 4]:
                total = sum(
                    (ws5.cell(row=r, column=ci).value or 0)
                    for r in range(3, sr5)
                    if isinstance(ws5.cell(row=r, column=ci).value, (int, float))
                )
                c = ws5.cell(row=sr5, column=ci, value=round(total, 2))
                c.number_format = '#,##0.00" h"'

        _autofit(ws5)

        # Liniendiagramm Trend
        if report.trend_data:
            n_t    = len(report.trend_data)
            line5  = LineChart()
            line5.title  = "Stundenentwicklung"
            line5.y_axis.title = "Stunden"
            line5.style  = 10
            line5.width  = 26
            line5.height = 14
            off_ref5 = Reference(ws5, min_col=2, min_row=2, max_row=2 + n_t)
            pnd_ref5 = Reference(ws5, min_col=3, min_row=2, max_row=2 + n_t)
            cat_ref5 = Reference(ws5, min_col=1, min_row=3, max_row=2 + n_t)
            line5.add_data(off_ref5, titles_from_data=True)
            line5.add_data(pnd_ref5, titles_from_data=True)
            line5.set_categories(cat_ref5)
            line5.series[0].graphicalProperties.line.solidFill = "1E3A5F"
            line5.series[0].graphicalProperties.line.width = 25000
            line5.series[0].smooth = True
            line5.series[1].graphicalProperties.line.solidFill = "C8A84B"
            line5.series[1].graphicalProperties.line.width = 20000
            line5.series[1].smooth = True
            chart_row5 = ws5.max_row + 2
            ws5.add_chart(line5, f"A{chart_row5}")

        # ── Streamen ──────────────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        data = buf.read()

    except HTTPException:
        raise
    except Exception as exc:
        logger.error(
            "Excel-Export V2 fehlgeschlagen (from=%s to=%s emps=%s locs=%s):\n%s",
            from_date, to_date, employee_ids, location_ids,
            traceback.format_exc(),
        )
        raise HTTPException(
            status_code=500,
            detail=f"Excel-Export fehlgeschlagen: {type(exc).__name__}: {exc}",
        )

    # Dateiname
    emp_ids_set = {e.employee_id for e in report.employee_summary}
    if len(employee_ids) == 1 and employee_ids[0] in emp_ids_set:
        raw   = next(
            (e.employee_name for e in report.employee_summary if e.employee_id == employee_ids[0]),
            "Mitarbeiter",
        )
        safe  = "".join(c if c.isalnum() else "_" for c in raw)
        fname = f"Arbeitszeitbericht_{safe}_{from_date}_{to_date}.xlsx"
    else:
        label = "Alle_Mitarbeiter" if not employee_ids else "Mehrere_Mitarbeiter"
        fname = f"Arbeitszeitbericht_{label}_{from_date}_{to_date}.xlsx"

    return StreamingResponse(
        iter([data]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
