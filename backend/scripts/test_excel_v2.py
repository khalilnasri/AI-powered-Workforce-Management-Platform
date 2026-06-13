"""
Lokaler Smoke-Test für den Excel-V2-Export.

Ausführung (aus dem backend/-Ordner):
    python scripts/test_excel_v2.py

Setzt USE_SQLITE=1 voraus (oder eine funktionierende PG-Verbindung über .env).
"""
from __future__ import annotations

import io
import os
import sys
import traceback

# .env laden, damit DATABASE_URL / USE_SQLITE gesetzt sind
from pathlib import Path
env_file = Path(__file__).parent.parent / ".env"
if env_file.exists():
    from dotenv import load_dotenv
    load_dotenv(env_file)

sys.path.insert(0, str(Path(__file__).parent.parent))

from datetime import datetime, timedelta, UTC
from zoneinfo import ZoneInfo

_BERLIN = ZoneInfo("Europe/Berlin")


def _run_excel_generation() -> bytes:
    """Führt die gesamte Excel-Erzeugung mit Dummy-Daten durch."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    # ── Dummy-Daten ──────────────────────────────────────────────────────────
    now_berlin = datetime.now(_BERLIN).replace(microsecond=0)
    ci_utc     = (now_berlin - timedelta(hours=8)).astimezone(UTC)
    co_utc     = now_berlin.astimezone(UTC)

    # Minimale Mock-Objekte
    class Row:
        employee_id   = 1
        employee_name = "Max Mustermann"
        date          = ci_utc.astimezone(_BERLIN).date().isoformat()
        weekday       = "Montag"
        location_name = "Hauptsitz Berlin"
        checkin_time  = ci_utc
        checkout_time = co_utc
        break_minutes = 0
        work_minutes  = 480
        duration_minutes = 480
        status        = "approved"

    class KPIs:
        total_hours    = 8.0
        official_hours = 8.0
        pending_hours  = 0.0
        total_shifts   = 1
        location_count = 1
        work_days      = 1

    class LocRow:
        location_name = "Hauptsitz Berlin"
        shift_count   = 1
        total_hours   = 8.0

    class EmpRow:
        employee_id    = 1
        employee_name  = "Max Mustermann"
        official_hours = 8.0
        pending_hours  = 0.0
        target_hours   = 160
        diff_hours     = -152.0
        shift_count    = 1
        work_days      = 1

    class PeriodSummary:
        total_hours    = 8.0
        official_hours = 8.0
        pending_hours  = 0.0
        target_hours   = 160
        diff_hours     = -152.0
        shift_count    = 1
        work_days      = 1

    class Report:
        kpis             = KPIs()
        sessions         = [Row()]
        location_summary = [LocRow()]
        employee_summary = [EmpRow()]
        period_summary   = PeriodSummary()

    report         = Report()
    emp_email_map  = {1: "max@example.com"}
    from_date      = "2026-06-01"
    to_date        = "2026-06-30"

    # ── Helpers (identisch mit dem Endpoint) ─────────────────────────────────
    def _to_berlin(dt):
        if dt is None:
            return None
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=UTC)
        return dt.astimezone(_BERLIN)

    def _fmt_dur(minutes):
        if minutes is None:
            return "—"
        minutes = int(round(minutes))
        sign = "-" if minutes < 0 else ""
        h, m = divmod(abs(minutes), 60)
        return f"{sign}{h}:{m:02d}"

    def _hdr(ws_sheet, cols, hdr_fill, hdr_font):
        for ci, c in enumerate(cols, 1):
            cell = ws_sheet.cell(row=1, column=ci, value=c)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _autofit(ws_sheet):
        try:
            for col in ws_sheet.columns:
                if not col:
                    continue
                w = max((len(str(cell.value or "")) for cell in col), default=0)
                ws_sheet.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 50)
        except Exception as e:
            print(f"  [WARN] _autofit: {e}")

    HDR_FILL    = PatternFill("solid", fgColor="FF1E3A5F")
    HDR_FONT    = Font(bold=True, color="FFFFFFFF", size=10)
    GREEN_FILL  = PatternFill("solid", fgColor="FFD1FAE5")
    RED_FILL    = PatternFill("solid", fgColor="FFFEE2E2")
    ORANGE_FILL = PatternFill("solid", fgColor="FFFEF3C7")
    BLUE_FILL   = PatternFill("solid", fgColor="FFDBEAFE")
    GRAY_FILL   = PatternFill("solid", fgColor="FFF8FAFC")
    BOLD        = Font(bold=True)

    STATUS_FILL = {"approved": GREEN_FILL, "corrected": BLUE_FILL,
                   "rejected": RED_FILL,   "pending":   ORANGE_FILL}
    STATUS_DE   = {"approved": "Genehmigt", "corrected": "Korrigiert",
                   "rejected": "Abgelehnt", "pending":   "Ausstehend"}

    ps        = report.period_summary
    k         = report.kpis
    multi_emp = len(report.employee_summary) > 1
    wb        = Workbook()

    # Sheet 1
    ws1 = wb.active
    ws1.title = "Zusammenfassung"
    emp_label = str(len(report.employee_summary)) if report.employee_summary else "Alle"
    for r in [
        ("Zeitraum", f"{from_date} – {to_date}"),
        ("Mitarbeiter", emp_label),
        ("", ""),
        ("KPI", "Wert"),
        ("Gesamtstunden (h)",       _fmt_dur(k.total_hours * 60)),
        ("Offizielle Stunden (h)",  _fmt_dur(ps.official_hours * 60)),
        ("Ausstehend (h)",          _fmt_dur(ps.pending_hours * 60)),
        ("Soll-Stunden (h)",        str(ps.target_hours)),
        ("Differenz (h)",           _fmt_dur(ps.diff_hours * 60)),
        ("Schichten",               str(k.total_shifts)),
        ("Arbeitstage",             str(k.work_days)),
        ("Standorte",               str(k.location_count)),
    ]:
        ws1.append(list(r))
    ws1["A4"].font = BOLD
    ws1["B4"].font = BOLD
    ws1.freeze_panes = "A2"
    _autofit(ws1)

    # Sheet 2
    ws2   = wb.create_sheet("Arbeitszeiten")
    cols2 = (["Mitarbeiter", "E-Mail"] if multi_emp else []) + [
        "Datum", "Wochentag", "Standort", "Check-In", "Check-Out",
        "Pause", "Arbeitszeit", "Status",
    ]
    _hdr(ws2, cols2, HDR_FILL, HDR_FONT)
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(cols2))}1"
    for row in report.sessions:
        ci_b = _to_berlin(row.checkin_time)
        co_b = _to_berlin(row.checkout_time)
        if ci_b is None:
            continue
        vals = (
            ([row.employee_name, emp_email_map.get(row.employee_id, "")] if multi_emp else [])
            + [
                ci_b.strftime("%d.%m.%Y"), row.weekday, row.location_name,
                ci_b.strftime("%H:%M"), co_b.strftime("%H:%M") if co_b else "—",
                "—", _fmt_dur(row.work_minutes), STATUS_DE.get(row.status, row.status),
            ]
        )
        ri = ws2.max_row + 1
        ws2.append(vals)
        fill = STATUS_FILL.get(row.status, GRAY_FILL)
        for ci in range(1, len(cols2) + 1):
            ws2.cell(row=ri, column=ci).fill = fill
    _autofit(ws2)

    # Sheet 3
    ws3   = wb.create_sheet("Standortauswertung")
    cols3 = ["Standort", "Anzahl Schichten", "Stunden"]
    _hdr(ws3, cols3, HDR_FILL, HDR_FONT)
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(cols3))}1"
    for loc in report.location_summary:
        ws3.append([loc.location_name, loc.shift_count, _fmt_dur(loc.total_hours * 60)])
    sum_row3 = ws3.max_row + 1
    ws3.append(["GESAMT", sum(l.shift_count for l in report.location_summary), _fmt_dur(k.total_hours * 60)])
    for ci in range(1, 4):
        ws3.cell(row=sum_row3, column=ci).font = BOLD
    _autofit(ws3)

    # Sheet 4
    ws4   = wb.create_sheet("Mitarbeiteruebersicht")
    cols4 = ["Mitarbeiter", "E-Mail", "Offizielle Stunden", "Ausstehend",
             "Schichten", "Arbeitstage", "Soll (h)", "Differenz (h)"]
    _hdr(ws4, cols4, HDR_FILL, HDR_FONT)
    ws4.freeze_panes = "A2"
    ws4.auto_filter.ref = f"A1:{get_column_letter(len(cols4))}1"
    for emp_row in report.employee_summary:
        diff = emp_row.diff_hours or 0.0
        ri   = ws4.max_row + 1
        ws4.append([
            emp_row.employee_name,
            emp_email_map.get(emp_row.employee_id, ""),
            _fmt_dur(emp_row.official_hours * 60),
            _fmt_dur(emp_row.pending_hours * 60),
            emp_row.shift_count,
            emp_row.work_days,
            str(emp_row.target_hours),
            _fmt_dur(diff * 60),
        ])
        fill = GREEN_FILL if diff >= 0 else RED_FILL
        for ci in [3, 8]:
            ws4.cell(row=ri, column=ci).fill = fill
    _autofit(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _run_empty_report():
    """Leerer Report – keine Sessions, keine Standorte, keine Mitarbeiter."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    def _fmt_dur(m):
        if m is None: return "—"
        m = int(round(m))
        sign = "-" if m < 0 else ""
        h, r = divmod(abs(m), 60)
        return f"{sign}{h}:{r:02d}"

    def _hdr(ws, cols, fill, font):
        for ci, c in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=c)
            cell.font = font; cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

    def _autofit(ws):
        try:
            for col in ws.columns:
                if not col: continue
                w = max((len(str(c.value or "")) for c in col), default=0)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(w+4, 50)
        except Exception: pass

    HDR_FILL = PatternFill("solid", fgColor="FF1E3A5F")
    HDR_FONT = Font(bold=True, color="FFFFFFFF", size=10)
    BOLD     = Font(bold=True)
    wb       = Workbook()

    ws1       = wb.active; ws1.title = "Zusammenfassung"
    for r in [("Zeitraum","2026-06-01 – 2026-06-30"),("Mitarbeiter","Alle"),("",""),
              ("KPI","Wert"),("Gesamtstunden (h)","0:00"),("Offizielle Stunden (h)","0:00"),
              ("Ausstehend (h)","0:00"),("Soll-Stunden (h)","0"),("Differenz (h)","0:00"),
              ("Schichten","0"),("Arbeitstage","0"),("Standorte","0")]:
        ws1.append(list(r))
    ws1["A4"].font = BOLD; ws1["B4"].font = BOLD
    _autofit(ws1)

    ws2   = wb.create_sheet("Arbeitszeiten")
    cols2 = ["Datum","Wochentag","Standort","Check-In","Check-Out","Pause","Arbeitszeit","Status"]
    _hdr(ws2, cols2, HDR_FILL, HDR_FONT)
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(cols2))}1"
    _autofit(ws2)

    ws3   = wb.create_sheet("Standortauswertung")
    cols3 = ["Standort","Anzahl Schichten","Stunden"]
    _hdr(ws3, cols3, HDR_FILL, HDR_FONT)
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(cols3))}1"
    ws3.append(["GESAMT", 0, "0:00"])
    ws3.cell(row=2, column=1).font = BOLD
    _autofit(ws3)

    ws4   = wb.create_sheet("Mitarbeiteruebersicht")
    cols4 = ["Mitarbeiter","E-Mail","Offizielle Stunden","Ausstehend",
             "Schichten","Arbeitstage","Soll (h)","Differenz (h)"]
    _hdr(ws4, cols4, HDR_FILL, HDR_FONT)
    ws4.auto_filter.ref = f"A1:{get_column_letter(len(cols4))}1"
    _autofit(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":
    print("=" * 60)
    print("Test 1: Normaler Report (1 Session, 1 Mitarbeiter)")
    try:
        data = _run_excel_generation()
        out  = Path("test_excel_v2_normal.xlsx")
        out.write_bytes(data)
        print(f"  OK – {len(data):,} Bytes → {out}")
    except Exception:
        print("  FEHLER:")
        traceback.print_exc()

    print()
    print("Test 2: Leerer Report (keine Sessions)")
    try:
        data = _run_empty_report()
        out  = Path("test_excel_v2_empty.xlsx")
        out.write_bytes(data)
        print(f"  OK – {len(data):,} Bytes → {out}")
    except Exception:
        print("  FEHLER:")
        traceback.print_exc()

    print()
    print("Test 3: Import-Check (openpyxl)")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        # PatternFill mit 8-Char ARGB (neue Format)
        pf = PatternFill("solid", fgColor="FF1E3A5F")
        # PatternFill mit 6-Char RGB (altes Format)
        pf2 = PatternFill("solid", fgColor="1E3A5F")
        print(f"  OK – openpyxl verfügbar, 8-char ARGB: {pf.fgColor.rgb!r}, 6-char RGB: {pf2.fgColor.rgb!r}")
    except Exception:
        print("  FEHLER:")
        traceback.print_exc()

    print()
    print("Fertig.")
